"""Test-Skript: Erzeugt eine absichtlich 'schlechte' Excel-Datei und prüft sie."""

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

import os
import sys

# Pfad für die Test-Datei
TEST_DIR = os.path.dirname(os.path.abspath(__file__))
TEST_FILE = os.path.join(TEST_DIR, "test_messy.xlsx")


def create_test_file():
    """Erstellt eine intentional problematische Excel-Datei."""
    wb = openpyxl.Workbook()

    # === Sheet 1: "Daten" – viele Probleme ===
    ws = wb.active
    ws.title = "Tabelle1"  # Kryptischer Name

    # Merged cells
    ws.merge_cells("A1:D1")
    ws["A1"] = "Projektübersicht Q4/2025"
    ws["A1"].font = Font(size=14, bold=True)

    ws.merge_cells("A2:D2")
    ws["A2"] = "Stand: aktuell"

    # Header erst in Zeile 4
    headers = ["ID", "Projektname", "Budget", "Status", "Verantwortlich"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)

    # Daten mit gemischten Typen und inkonsistenten IDs
    data = [
        ("PRJ-001", "Website Relaunch", 150000, "aktiv", "Müller"),
        ("PRJ-02", "ERP Migration", 500000, "aktiv", "Schmidt"),  # ID Inkonsistenz!
        ("PRJ-003", "Mobile App", "noch offen", "geplant", "Weber"),  # Budget = Text!
        ("PRJ-004", "Data Warehouse", 300000, "aktiv", "Fischer"),
        (None, None, None, None, None),  # Leere Trennzeile
        ("PRJ-006", "CRM Update", 80000, "abgeschlossen", "Wagner"),  # Lücke in ID!
        ("PRJ-007", "Security Audit", 45000, "aktiv", "Bauer"),
        ("PRJ-007", "Security Audit v2", 48000, "aktiv", "Bauer"),  # Doppelte ID!
    ]
    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Farbcodes ohne Legende
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

    ws.cell(row=5, column=4).fill = green_fill
    ws.cell(row=6, column=4).fill = green_fill
    ws.cell(row=7, column=4).fill = yellow_fill
    ws.cell(row=8, column=4).fill = green_fill
    ws.cell(row=10, column=4).fill = blue_fill
    ws.cell(row=11, column=4).fill = green_fill
    ws.cell(row=12, column=4).fill = red_fill

    # Kommentar mit Geschäftslogik
    ws.cell(row=7, column=3).comment = Comment(
        "ACHTUNG: Dieses Budget NICHT ändern! Muss immer mit Vorstand abgestimmt werden.",
        "AdminUser"
    )

    # Hardcodierte Validierungsliste
    dv = DataValidation(type="list", formula1='"aktiv,geplant,abgeschlossen,pausiert,storniert"')
    dv.add("D5:D100")
    ws.add_data_validation(dv)

    # === Sheet 2: Versteckte Hilfstabelle ===
    ws2 = wb.create_sheet("Hilfstabelle")
    ws2.sheet_state = 'hidden'
    ws2["A1"] = "MWSt-Satz"
    ws2["B1"] = 0.20
    ws2["A2"] = "Rabatt"
    ws2["B2"] = 0.15

    # === Sheet 3: "Kopie von Tabelle1 (2)" ===
    ws3 = wb.create_sheet("Kopie von Tabelle1 (2)")
    ws3["A1"] = "Alte Version - nicht verwenden"

    # === Sheet 4: Formeln mit Problemen ===
    ws4 = wb.create_sheet("Berechnungen")

    # Volatile Funktionen
    ws4["A1"] = "Heute"
    ws4["B1"] = "=HEUTE()"
    ws4["A2"] = "Berechnung"
    ws4["B2"] = "=SUMME($C$5:$C$100)*1.2*0.85"  # Magic Numbers + absolute Bezüge
    ws4["A3"] = "Verweis"
    ws4["B3"] = "=SVERWEIS(A5,Tabelle1!A:E,3,FALSCH)"

    # Viele Lookups simulieren
    for i in range(5, 60):
        ws4.cell(row=i, column=1, value=f"=SVERWEIS(A{i},Tabelle1!A:E,{(i % 4)+1},FALSCH)")
        ws4.cell(row=i, column=2, value=f"=WENN(A{i}>1000,A{i}*0.19,A{i}*0.07)")

    # Versteckte Zeilen
    for r in range(20, 30):
        ws4.row_dimensions[r].hidden = True

    # Versteckte Spalten
    ws4.column_dimensions['D'].hidden = True
    ws4.column_dimensions['E'].hidden = True
    ws4.column_dimensions['F'].hidden = True

    # Bedingte Formatierungen
    from openpyxl.formatting.rule import CellIsRule
    for i in range(15):
        ws4.conditional_formatting.add(
            f"A{i+1}:D{i+1}",
            CellIsRule(operator='greaterThan', formula=['1000'],
                       fill=PatternFill(bgColor="FFC7CE"))
        )

    # Blattschutz
    ws4.protection.sheet = True

    # Speichern
    wb.save(TEST_FILE)
    print(f"✅ Test-Datei erstellt: {TEST_FILE}")
    return TEST_FILE


def run_check(file_path: str):
    """Führt den Health Check aus."""
    from excel_checker.engine import analyze
    from excel_checker.report import generate_html

    print(f"\n🔍 Analysiere: {file_path}")
    report = analyze(file_path)

    # Score
    score = report.health_score
    print(f"\n{'='*50}")
    print(f"  Health-Score: {score}/100")
    print(f"  Findings: {len(report.findings)}")
    print(f"  Empfehlungen: {len(report.recommendations)}")
    print(f"{'='*50}")

    # Findings ausgeben
    for f in report.findings:
        sev = {"critical": "🔴", "error": "🟠", "warning": "🟡", "info": "🔵"}
        sym = sev.get(f.severity.value, "⚪")
        print(f"\n{sym} [{f.rule_id}] {f.message}")
        if f.sheet:
            print(f"   Sheet: {f.sheet}")
        if f.detail:
            print(f"   Details: {f.detail}")
        if f.suggestion:
            print(f"   💡 {f.suggestion}")

    # Empfehlungen
    print(f"\n{'='*50}")
    print("🎯 STRATEGISCHE EMPFEHLUNGEN")
    print(f"{'='*50}")
    for rec in report.recommendations:
        print(f"\n📌 Priorität {rec.priority}: {rec.title}")
        print(f"   Typ: {rec.rec_type.value}")
        print(f"   Grund: {rec.reason}")
        print(f"   Maßnahme: {rec.action}")

    # HTML-Report
    html_path = os.path.splitext(file_path)[0] + "_health_report.html"
    html = generate_html(report)
    with open(html_path, "w", encoding="utf-8") as fh:
        fh.write(html)
    print(f"\n📄 HTML-Report generiert: {html_path}")

    return report


if __name__ == "__main__":
    test_file = create_test_file()
    run_check(test_file)
