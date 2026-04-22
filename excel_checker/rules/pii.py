"""PII-Regeln: Email, IBAN, SV-Nummer, Telefon (Phase 2 A.2).

Erkennt sensible Daten in String-Zellen. Findings starten als WARNING
(nicht CRITICAL), um False-Positives bei z. B. Produktcodes zu tolerieren;
Tuning nach Dogfood.
"""

from __future__ import annotations

import re
from typing import Iterable, List, Optional

import openpyxl
from openpyxl.utils import get_column_letter

from excel_checker.models import Category, Finding, Severity
from excel_checker.rules._patterns import (
    EMAIL_PATTERN, IBAN_PATTERN, PHONE_PATTERN, SVNR_AT_PATTERN,
    iban_checksum_ok, svnr_at_checksum_ok,
)
from excel_checker.rules._sampling import SampleMode, iter_sampled_rows
from excel_checker.rules.base import BaseRule


_MAX_EXAMPLES = 5


def _iter_string_cells(
    ws, sample_mode: Optional[SampleMode],
) -> Iterable[tuple[int, int, str]]:
    """Liefert Tupel (row_idx, col_idx, text) für alle nicht-leeren String-Zellen."""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return
    for row_idx, row in iter_sampled_rows(ws, max_row, max_col, sample_mode):
        for col_idx, cell in enumerate(row, start=1):
            val = cell.value
            if isinstance(val, str) and val.strip():
                yield row_idx, col_idx, val


def _emit_finding(rule_id: str, rule_name: str, kind_label: str,
                  ws_title: str, col_idx: int, row_idx: int, match_text: str,
                  severity: Severity, score_penalty: int,
                  sample_note: Optional[str]) -> Finding:
    cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
    return Finding(
        rule_id=rule_id,
        rule_name=rule_name,
        category=Category.STRUCTURE,  # Audience-Mapping läuft über Prefix PII-
        severity=severity,
        message=f"{kind_label} in Zelle {cell_ref} erkannt",
        sheet=ws_title,
        cell=cell_ref,
        detail=f"Treffer: {match_text[:60]}{'…' if len(match_text) > 60 else ''}",
        suggestion=(
            "Sensible Daten gehören nicht in offene Excel-Dateien. "
            "Entferne den Eintrag, pseudonymisiere ihn, oder verlege ihn in ein geschütztes System."
        ),
        score_penalty=score_penalty,
        sample_note=sample_note,
    )


class _BasePiiRule(BaseRule):
    """Gemeinsames Grundgerüst für alle PII-Rules."""

    needs_styles = False
    scales_with_cells = True
    supports_sampling = True

    # Subklassen überschreiben diese:
    _kind_label: str = "PII"
    _pattern: re.Pattern | None = None
    _severity: Severity = Severity.WARNING
    _score_penalty: int = 4

    def _find_matches(self, text: str) -> list[str]:
        """Subklassen können Prüfsummen-Validierung zusätzlich anwenden."""
        if self._pattern is None:
            return []
        return [m.group(0) for m in self._pattern.finditer(text)]

    def check(self, workbook: openpyxl.Workbook, file_path: str,
              progress_callback=None, sample_mode=None) -> List[Finding]:
        findings: List[Finding] = []
        note = sample_mode.disclosure_de() if sample_mode is not None else None
        max_findings_per_sheet = 20  # Flood-Schutz
        for ws in workbook.worksheets:
            emitted = 0
            for row_idx, col_idx, text in _iter_string_cells(ws, sample_mode):
                if emitted >= max_findings_per_sheet:
                    break
                matches = self._find_matches(text)
                if not matches:
                    continue
                findings.append(_emit_finding(
                    self.rule_id, self.rule_name, self._kind_label,
                    ws.title, col_idx, row_idx, matches[0],
                    self._severity, self._score_penalty, note,
                ))
                emitted += 1
        return findings


class EmailPiiRule(_BasePiiRule):
    rule_id = "PII-001"
    rule_name = "E-Mail-Adresse erkannt"
    _kind_label = "E-Mail-Adresse"
    _pattern = EMAIL_PATTERN
    _severity = Severity.WARNING
    _score_penalty = 3


class IbanPiiRule(_BasePiiRule):
    rule_id = "PII-002"
    rule_name = "IBAN erkannt"
    _kind_label = "IBAN (Verdacht)"
    _pattern = IBAN_PATTERN
    _severity = Severity.WARNING
    _score_penalty = 6

    def _find_matches(self, text: str) -> list[str]:
        if self._pattern is None:
            return []
        hits = []
        for m in self._pattern.finditer(text):
            candidate = m.group(0)
            if iban_checksum_ok(candidate):
                hits.append(candidate)
        return hits


class SvnrPiiRule(_BasePiiRule):
    rule_id = "PII-003"
    rule_name = "Sozialversicherungsnummer erkannt"
    _kind_label = "Sozialversicherungsnummer (AT)"
    _pattern = SVNR_AT_PATTERN
    _severity = Severity.WARNING
    _score_penalty = 8

    def _find_matches(self, text: str) -> list[str]:
        if self._pattern is None:
            return []
        hits = []
        for m in self._pattern.finditer(text):
            if svnr_at_checksum_ok(m):
                hits.append(m.group(0))
        return hits


class PhonePiiRule(_BasePiiRule):
    rule_id = "PII-004"
    rule_name = "Telefonnummer erkannt"
    _kind_label = "Telefonnummer (Verdacht)"
    _pattern = PHONE_PATTERN
    _severity = Severity.INFO  # konservativ — Phone ist am anfälligsten für FP
    _score_penalty = 2
