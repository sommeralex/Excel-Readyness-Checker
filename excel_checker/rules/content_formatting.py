"""Content-Formatting-Regeln (Phase 2 A.2).

Erkennt Whitespace-, Zero-Width-, Number-as-Text-, Datumsformat- und
Encoding-Mojibake-Probleme in Zell-Werten.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from typing import List, Optional

import openpyxl
from openpyxl.utils import get_column_letter

from excel_checker.models import Category, Finding, Severity
from excel_checker.rules._patterns import (
    contains_mojibake, contains_zero_width, date_format_key,
    has_leading_or_trailing_whitespace, looks_like_number_text,
)
from excel_checker.rules._sampling import SampleMode, iter_sampled_rows
from excel_checker.rules.base import BaseRule


_MAX_EXAMPLES = 3


def _make_finding(rule_id: str, rule_name: str, message: str, ws_title: str,
                  cell_ref: Optional[str], detail: str, suggestion: str,
                  severity: Severity, penalty: int,
                  sample_note: Optional[str]) -> Finding:
    return Finding(
        rule_id=rule_id,
        rule_name=rule_name,
        category=Category.STRUCTURE,
        severity=severity,
        message=message,
        sheet=ws_title,
        cell=cell_ref,
        detail=detail,
        suggestion=suggestion,
        score_penalty=penalty,
        sample_note=sample_note,
    )


class WhitespaceAnomalyRule(BaseRule):
    """Leading/Trailing-Whitespace in Zellwerten — brechen JOINs und Duplikat-Checks."""

    rule_id = "CNT-001"
    rule_name = "Unsichtbare Leerzeichen am Anfang/Ende"
    scales_with_cells = True
    supports_sampling = True

    def check(self, workbook, file_path, progress_callback=None, sample_mode=None):
        findings: List[Finding] = []
        note = sample_mode.disclosure_de() if sample_mode else None
        for ws in workbook.worksheets:
            examples: list[str] = []
            count = 0
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            if max_row == 0 or max_col == 0:
                continue
            for row_idx, row in iter_sampled_rows(ws, max_row, max_col, sample_mode):
                for col_idx, cell in enumerate(row, start=1):
                    val = cell.value
                    if isinstance(val, str) and has_leading_or_trailing_whitespace(val):
                        count += 1
                        if len(examples) < _MAX_EXAMPLES:
                            examples.append(f"{get_column_letter(col_idx)}{row_idx}: {val!r}")
            if count > 0:
                findings.append(_make_finding(
                    self.rule_id, self.rule_name,
                    f"{count} Zellen haben unsichtbare Leerzeichen am Anfang/Ende",
                    ws.title, None,
                    "Beispiele: " + " | ".join(examples),
                    "Zellen mit TRIM() oder per Suchen & Ersetzen bereinigen — "
                    "sonst schlagen Vergleiche und JOINs fehl.",
                    Severity.WARNING, min(15, count // 2 + 2), note,
                ))
        return findings


class ZeroWidthCharRule(BaseRule):
    """Zero-Width-Zeichen in Zellwerten — unsichtbar, aber vergleichs-relevant."""

    rule_id = "CNT-002"
    rule_name = "Unsichtbare Steuerzeichen (Zero-Width)"
    scales_with_cells = True
    supports_sampling = True

    def check(self, workbook, file_path, progress_callback=None, sample_mode=None):
        findings: List[Finding] = []
        note = sample_mode.disclosure_de() if sample_mode else None
        for ws in workbook.worksheets:
            examples: list[str] = []
            count = 0
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            if max_row == 0 or max_col == 0:
                continue
            for row_idx, row in iter_sampled_rows(ws, max_row, max_col, sample_mode):
                for col_idx, cell in enumerate(row, start=1):
                    val = cell.value
                    if isinstance(val, str) and contains_zero_width(val):
                        count += 1
                        if len(examples) < _MAX_EXAMPLES:
                            examples.append(f"{get_column_letter(col_idx)}{row_idx}")
            if count > 0:
                findings.append(_make_finding(
                    self.rule_id, self.rule_name,
                    f"{count} Zellen enthalten unsichtbare Steuerzeichen",
                    ws.title, None,
                    "Betroffene Zellen: " + ", ".join(examples),
                    "Zero-Width-Zeichen (häufig aus Web-Copy-Paste) entfernen — "
                    "sie brechen Textvergleiche lautlos.",
                    Severity.WARNING, min(10, count + 1), note,
                ))
        return findings


class NumberAsTextRule(BaseRule):
    """Zahlen als Strings — blockieren arithmetische Auswertung & Migration."""

    rule_id = "CNT-003"
    rule_name = "Zahlen als Text gespeichert"
    scales_with_cells = True
    supports_sampling = True

    def check(self, workbook, file_path, progress_callback=None, sample_mode=None):
        findings: List[Finding] = []
        note = sample_mode.disclosure_de() if sample_mode else None
        for ws in workbook.worksheets:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            if max_row == 0 or max_col == 0:
                continue
            # Pro Spalte: wenn ≥3 Werte Zahl-als-Text, aggregieren.
            cols_with_hits: dict[int, list[tuple[int, str]]] = defaultdict(list)
            for row_idx, row in iter_sampled_rows(ws, max_row, max_col, sample_mode):
                if row_idx == 1:
                    continue  # Header überspringen
                for col_idx, cell in enumerate(row, start=1):
                    val = cell.value
                    if isinstance(val, str) and looks_like_number_text(val):
                        cols_with_hits[col_idx].append((row_idx, val))
            for col_idx, hits in cols_with_hits.items():
                if len(hits) < 3:
                    continue
                col_letter = get_column_letter(col_idx)
                ex = ", ".join(f"{col_letter}{r}={v!r}" for r, v in hits[:_MAX_EXAMPLES])
                findings.append(_make_finding(
                    self.rule_id, self.rule_name,
                    f"Spalte {col_letter}: {len(hits)} Zahlen sind als Text gespeichert",
                    ws.title, f"{col_letter}{hits[0][0]}",
                    "Beispiele: " + ex,
                    "Spalte markieren, über 'Daten → Text in Spalten' in Zahl umwandeln "
                    "oder per Formel =WERT(A1) konvertieren.",
                    Severity.WARNING, min(12, len(hits) // 3 + 2), note,
                ))
        return findings


class DateFormatInconsistencyRule(BaseRule):
    """Unterschiedliche Datumsformate innerhalb einer Spalte."""

    rule_id = "CNT-004"
    rule_name = "Uneinheitliche Datumsformate in einer Spalte"
    scales_with_cells = True
    supports_sampling = True

    def check(self, workbook, file_path, progress_callback=None, sample_mode=None):
        findings: List[Finding] = []
        note = sample_mode.disclosure_de() if sample_mode else None
        for ws in workbook.worksheets:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            if max_row == 0 or max_col == 0:
                continue
            col_formats: dict[int, Counter] = defaultdict(Counter)
            col_samples: dict[int, dict[str, str]] = defaultdict(dict)
            for row_idx, row in iter_sampled_rows(ws, max_row, max_col, sample_mode):
                if row_idx == 1:
                    continue
                for col_idx, cell in enumerate(row, start=1):
                    val = cell.value
                    if not isinstance(val, str):
                        continue
                    key = date_format_key(val)
                    if key is None:
                        continue
                    col_formats[col_idx][key] += 1
                    col_samples[col_idx].setdefault(key, f"{get_column_letter(col_idx)}{row_idx}: {val!r}")
            for col_idx, counts in col_formats.items():
                if len(counts) < 2:
                    continue
                col_letter = get_column_letter(col_idx)
                distribution = ", ".join(f"{k}×{v}" for k, v in counts.most_common())
                example_lines = " | ".join(col_samples[col_idx].values())
                findings.append(_make_finding(
                    self.rule_id, self.rule_name,
                    f"Spalte {col_letter}: {len(counts)} verschiedene Datumsformate",
                    ws.title, None,
                    f"Verteilung: {distribution}. Beispiele: {example_lines}",
                    "Ein einheitliches Format für die ganze Spalte festlegen "
                    "(idealerweise ISO YYYY-MM-DD).",
                    Severity.WARNING, 4, note,
                ))
        return findings


class EncodingMojibakeRule(BaseRule):
    """Doppelt-kodierte UTF-8-Sequenzen (Ã¤, Ã¼ etc.) — Encoding-Bruch."""

    rule_id = "CNT-005"
    rule_name = "Beschädigtes Encoding (Mojibake)"
    scales_with_cells = True
    supports_sampling = True

    def check(self, workbook, file_path, progress_callback=None, sample_mode=None):
        findings: List[Finding] = []
        note = sample_mode.disclosure_de() if sample_mode else None
        for ws in workbook.worksheets:
            examples: list[str] = []
            count = 0
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            if max_row == 0 or max_col == 0:
                continue
            for row_idx, row in iter_sampled_rows(ws, max_row, max_col, sample_mode):
                for col_idx, cell in enumerate(row, start=1):
                    val = cell.value
                    if isinstance(val, str) and contains_mojibake(val):
                        count += 1
                        if len(examples) < _MAX_EXAMPLES:
                            examples.append(f"{get_column_letter(col_idx)}{row_idx}: {val!r}")
            if count > 0:
                findings.append(_make_finding(
                    self.rule_id, self.rule_name,
                    f"{count} Zellen mit beschädigtem Encoding (Ã¤, Ã¼ statt ä, ü)",
                    ws.title, None,
                    "Beispiele: " + " | ".join(examples),
                    "Quelldaten mit korrektem Encoding re-importieren "
                    "(Excel: 'Aus Text/CSV' → UTF-8 auswählen).",
                    Severity.ERROR, min(15, count + 3), note,
                ))
        return findings
