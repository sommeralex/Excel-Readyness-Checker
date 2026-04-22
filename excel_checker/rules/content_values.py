"""Content-Values-Regeln (Phase 2 A.2).

Erkennt sich widersprechende Schreibweisen desselben logischen Werts
(Free-Text-in-Enum) — typisch für Status-Spalten wie
"aktiv" / "Aktiv" / "ACTIVE" / "active ".
"""

from __future__ import annotations

from collections import defaultdict
from typing import List

import openpyxl
from openpyxl.utils import get_column_letter

from excel_checker.models import Category, Finding, Severity
from excel_checker.rules._patterns import group_by_normalized
from excel_checker.rules._sampling import iter_sampled_rows
from excel_checker.rules.base import BaseRule


class FreeTextInEnumRule(BaseRule):
    """Spaltenwerte, die nach Normalisierung gleich sind, aber unterschiedlich
    geschrieben wurden — klassisches Enum-Chaos.
    """

    rule_id = "CNT-006"
    rule_name = "Inkonsistente Schreibweisen in einer Kategorie-Spalte"
    scales_with_cells = True
    supports_sampling = True

    # Nur auf Spalten mit begrenztem Wertebereich anwenden.
    _max_cardinality_for_enum = 30
    _min_duplicates_per_group = 2

    def check(self, workbook, file_path, progress_callback=None, sample_mode=None):
        findings: List[Finding] = []
        note = sample_mode.disclosure_de() if sample_mode else None
        for ws in workbook.worksheets:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            if max_row == 0 or max_col == 0:
                continue
            values_per_col: dict[int, list[str]] = defaultdict(list)
            for row_idx, row in iter_sampled_rows(ws, max_row, max_col, sample_mode):
                if row_idx == 1:
                    continue
                for col_idx, cell in enumerate(row, start=1):
                    val = cell.value
                    if isinstance(val, str) and val.strip():
                        values_per_col[col_idx].append(val)

            for col_idx, values in values_per_col.items():
                if len(values) < 5:
                    continue
                # Nur "kategorische" Spalten: niedrige Unique-Cardinality
                unique_count = len({v.strip() for v in values})
                if unique_count > self._max_cardinality_for_enum:
                    continue
                groups = group_by_normalized(values)
                # Relevante Gruppen: mindestens zwei unterschiedliche Schreibweisen,
                # jede davon mindestens mehrmals vorhanden.
                problems = []
                for norm, variants in groups.items():
                    from collections import Counter as _Counter
                    counter = _Counter(v.strip() for v in variants)
                    if len(counter) < 2:
                        continue
                    if max(counter.values()) < self._min_duplicates_per_group:
                        continue
                    problems.append((norm, counter))
                if not problems:
                    continue
                col_letter = get_column_letter(col_idx)
                detail_lines = []
                for norm, counter in problems[:3]:
                    mix = ", ".join(f"{repr(k)}×{v}" for k, v in counter.most_common())
                    detail_lines.append(f"'{norm}': {mix}")
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=self.rule_name,
                    category=Category.STRUCTURE,
                    severity=Severity.WARNING,
                    message=(
                        f"Spalte {col_letter}: {len(problems)} Kategorien "
                        "haben unterschiedliche Schreibweisen"
                    ),
                    sheet=ws.title,
                    detail=" | ".join(detail_lines),
                    suggestion=(
                        "Werte auf einheitliche Schreibweise vereinheitlichen "
                        "(z. B. alle kleinschreiben, Dropdown-Validierung ergänzen)."
                    ),
                    score_penalty=min(10, len(problems) * 2),
                    sample_note=note,
                ))
        return findings
