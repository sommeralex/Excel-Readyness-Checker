"""Regeln für implizites Wissen – das unsichtbare Risiko in Excel-Dateien."""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from typing import List, Optional, Tuple

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

from excel_checker.i18n import _
from excel_checker.models import Category, Finding, Severity
from excel_checker.rules._sampling import iter_window_rows
from excel_checker.rules.base import BaseRule


def _color_to_hex(color) -> Optional[str]:
    """Konvertiert openpyxl Color-Objekte in Hex-String."""
    if color is None:
        return None
    if hasattr(color, 'rgb') and color.rgb and color.rgb != '00000000':
        rgb = str(color.rgb)
        # openpyxl gibt manchmal ARGB zurück (8 Zeichen)
        if len(rgb) == 8:
            return f"#{rgb[2:]}"
        elif len(rgb) == 6:
            return f"#{rgb}"
    if hasattr(color, 'theme') and color.theme is not None:
        return f"Theme:{color.theme}"
    if hasattr(color, 'indexed') and color.indexed is not None:
        return f"Indexed:{color.indexed}"
    return None


def _is_default_color(hex_color: Optional[str]) -> bool:
    """Prüft ob eine Farbe die Standard-Farbe ist (weiß/schwarz/keine)."""
    if hex_color is None:
        return True
    defaults = {"#FFFFFF", "#000000", "#FFFFFF00", "Theme:0", "Theme:1",
                "Indexed:64", "Indexed:0"}
    return hex_color.upper() in defaults


class UndocumentedColorCodesRule(BaseRule):
    """Erkennt Farbcodes, die nicht durch eine Legende erklärt werden."""

    rule_id = "IMP-001"
    rule_name = _("IMP-001.name")
    needs_styles = True
    scales_with_cells = True
    supports_sampling = True

    # Begriffe die auf eine Legende hindeuten
    LEGEND_KEYWORDS = {
        "legende", "legend", "farbcode", "color code", "farbe", "bedeutung",
        "erklärung", "status", "ampel", "kennzeichnung",
    }

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        has_legend = self._has_legend_sheet(workbook)

        for ws in workbook.worksheets:
            if ws.max_row is None or ws.max_row < 3:
                continue
            max_row = min(ws.max_row, 5000)
            max_col = min(ws.max_column or 1, 200)

            bg_colors: Counter = Counter()
            font_colors: Counter = Counter()
            color_examples: dict[str, list] = defaultdict(list)

            for row_idx, row in iter_window_rows(ws, max_row, max_col):
                for col_pos, cell in enumerate(row):
                    col_idx = col_pos + 1

                    # Hintergrundfarbe (GradientFill hat kein fgColor)
                    try:
                        if (cell.fill
                                and hasattr(cell.fill, 'fgColor')
                                and cell.fill.fgColor):
                            bg = _color_to_hex(cell.fill.fgColor)
                            if bg and not _is_default_color(bg):
                                bg_colors[bg] += 1
                                if len(color_examples[bg]) < 3:
                                    col_letter = get_column_letter(col_idx)
                                    color_examples[bg].append(f"{col_letter}{row_idx}")
                    except (AttributeError, TypeError):
                        pass

                    # Schriftfarbe
                    if cell.font and cell.font.color:
                        fc = _color_to_hex(cell.font.color)
                        if fc and not _is_default_color(fc):
                            font_colors[fc] += 1

            unique_bg = len(bg_colors)
            unique_font = len(font_colors)
            total_colored = sum(bg_colors.values())

            if unique_bg >= 3 and total_colored > 10 and not has_legend:
                color_summary = ", ".join(
                    f"{color} ({count}x, z.B. {', '.join(color_examples.get(color, []))})"
                    for color, count in bg_colors.most_common(5)
                )
                # Sammle Beispielzellen für die wichtigsten Farben
                example_cells = []
                for color, _ in bg_colors.most_common(3):
                    example_cells.extend(color_examples.get(color, [])[:2])
                cell_str = ", ".join(example_cells) if example_cells else None
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("IMP-001.name"),
                    category=Category.STRUCTURE,
                    severity=Severity.WARNING,
                    message=_("IMP-001.msg.many", unique=unique_bg, total=total_colored),
                    sheet=ws.title,
                    cell=cell_str,
                    detail=_("IMP-001.detail", colors=color_summary),
                    suggestion=_("IMP-001.tip.many"),
                    score_penalty=min(10, unique_bg * 2),
                ))
            elif unique_bg >= 2 and total_colored > 5 and not has_legend:
                # Auch hier Beispielzellen angeben
                example_cells = []
                for color, _ in bg_colors.most_common(2):
                    example_cells.extend(color_examples.get(color, [])[:2])
                cell_str = ", ".join(example_cells) if example_cells else None
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("IMP-001.name"),
                    category=Category.STRUCTURE,
                    severity=Severity.INFO,
                    message=_("IMP-001.msg.some", unique=unique_bg),
                    sheet=ws.title,
                    cell=cell_str,
                    suggestion=_("IMP-001.tip.some"),
                    score_penalty=2,
                ))

            if unique_font >= 3 and not has_legend:
                # Für Schriftfarben keine Zellenbeispiele verfügbar, daher cell leer lassen
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("IMP-001.name"),
                    category=Category.STRUCTURE,
                    severity=Severity.INFO,
                    message=_("IMP-001.msg.font", unique=unique_font),
                    sheet=ws.title,
                    suggestion=_("IMP-001.tip.font"),
                    score_penalty=2,
                ))

        return findings

    def _has_legend_sheet(self, workbook: openpyxl.Workbook) -> bool:
        """Prüft ob es ein Sheet oder einen Bereich gibt, der als Legende dient."""
        for ws in workbook.worksheets:
            # Sheet-Name enthält Legende-Keyword?
            if any(kw in ws.title.lower() for kw in self.LEGEND_KEYWORDS):
                return True
            # Suche in den ersten 20 Zeilen nach Legende-Begriffen
            max_row = min(ws.max_row or 1, 20)
            max_col = min(ws.max_column or 1, 20)
            for _row_idx, row in iter_window_rows(ws, max_row, max_col):
                for cell in row:
                    if isinstance(cell.value, str):
                        if any(kw in cell.value.lower()
                               for kw in self.LEGEND_KEYWORDS):
                            return True
        return False


class HiddenSheetsRule(BaseRule):
    """Erkennt versteckte Tabellenblätter."""

    rule_id = "IMP-002"
    rule_name = _("IMP-002.name")

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        hidden = []
        very_hidden = []

        for ws in workbook.worksheets:
            if ws.sheet_state == 'hidden':
                hidden.append(ws.title)
            elif ws.sheet_state == 'veryHidden':
                very_hidden.append(ws.title)

        if very_hidden:
            findings.append(Finding(
                rule_id=self.rule_id,
                rule_name=_("IMP-002.name"),
                category=Category.STRUCTURE,
                severity=Severity.ERROR,
                message=_("IMP-002.msg.very_hidden", count=len(very_hidden)),
                detail=_("IMP-002.detail", sheets=', '.join(very_hidden)),
                suggestion=_("IMP-002.tip.very_hidden"),
                score_penalty=min(10, len(very_hidden) * 5),
            ))

        if hidden:
            findings.append(Finding(
                rule_id=self.rule_id,
                rule_name=_("IMP-002.name"),
                category=Category.STRUCTURE,
                severity=Severity.WARNING,
                message=_("IMP-002.msg.hidden", count=len(hidden)),
                detail=_("IMP-002.detail", sheets=', '.join(hidden)),
                suggestion=_("IMP-002.tip.hidden"),
                score_penalty=min(5, len(hidden) * 2),
            ))

        return findings


class HiddenRowsColumnsRule(BaseRule):
    """Erkennt versteckte Zeilen und Spalten."""

    rule_id = "IMP-003"
    rule_name = _("IMP-003.name")
    needs_styles = True  # row_dimensions/column_dimensions sind im read_only-Modus leer

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            if ws.max_row is None:
                continue

            hidden_rows = 0
            hidden_cols = 0

            # Versteckte Zeilen
            for row_idx in range(1, min(ws.max_row + 1, 5000)):
                rd = ws.row_dimensions.get(row_idx)
                if rd and rd.hidden:
                    hidden_rows += 1

            # Versteckte Spalten
            for col_idx in range(1, min((ws.max_column or 1) + 1, 200)):
                col_letter = get_column_letter(col_idx)
                cd = ws.column_dimensions.get(col_letter)
                if cd and cd.hidden:
                    hidden_cols += 1

            if hidden_rows > 5 or hidden_cols > 2:
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("IMP-003.name"),
                    category=Category.STRUCTURE,
                    severity=Severity.WARNING,
                    message=_("IMP-003.msg", rows=hidden_rows, cols=hidden_cols),
                    sheet=ws.title,
                    suggestion=_("IMP-003.tip"),
                    score_penalty=min(5, (hidden_rows // 10 + hidden_cols)),
                ))

        return findings


class ConditionalFormattingRule(BaseRule):
    """Erkennt bedingte Formatierungen als implizite Geschäftslogik."""

    rule_id = "IMP-004"
    rule_name = "Bedingte Formatierungen (Implizite Logik)"
    needs_styles = True  # ws.conditional_formatting ist im read_only-Modus nicht verfügbar

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            cf_rules = ws.conditional_formatting
            rule_count = len(list(cf_rules))

            if rule_count > 10:
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=self.rule_name,
                    category=Category.STRUCTURE,
                    severity=Severity.WARNING if rule_count < 30 else Severity.ERROR,
                    message=(
                        f"Geschäftslogik in Formatierung versteckt: "
                        f"{rule_count} bedingte Formatierungsregeln – diese "
                        f"Regeln sind schwer wartbar und für andere unsichtbar."
                    ),
                    sheet=ws.title,
                    suggestion=(
                        "Tipp: Bedingte Formatierungen bilden oft wichtige "
                        "Geschäftsregeln ab (z.B. 'rot wenn überfällig'). "
                        "Besser: Eine eigene Status-Spalte mit Formeln anlegen – "
                        "dann ist die Logik sichtbar, dokumentierbar und auswertbar."
                    ),
                    score_penalty=min(8, rule_count // 5),
                ))

        return findings


class MagicNumbersRule(BaseRule):
    """Erkennt hartcodierte Zahlen in Formeln ohne benannte Bereiche."""

    rule_id = "IMP-005"
    rule_name = "Hartcodierte Werte in Formeln"
    scales_with_cells = True
    supports_sampling = True

    # Harmlose Zahlen (0, 1, 2, 100 usw.)
    HARMLESS = {0, 1, 2, 10, 100, 1000, 12, 24, 365, 360}

    MAGIC_PATTERN = re.compile(r'(?<![A-Z$:])(\d+\.?\d+)(?![A-Z\d:$])', re.IGNORECASE)

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            if ws.max_row is None or ws.max_row < 2:
                continue
            max_row = min(ws.max_row, 3000)
            max_col = min(ws.max_column or 1, 200)

            magic_values: Counter = Counter()
            magic_examples: dict[str, list] = defaultdict(list)

            for row_idx, row in iter_window_rows(ws, max_row, max_col):
                for col_pos, cell in enumerate(row):
                    if cell.data_type != 'f' or not isinstance(cell.value, str):
                        continue
                    col_idx = col_pos + 1
                    formula = cell.value
                    numbers = self.MAGIC_PATTERN.findall(formula)
                    for num_str in numbers:
                        try:
                            num = float(num_str)
                        except ValueError:
                            continue
                        if num not in self.HARMLESS and num > 2:
                            magic_values[num_str] += 1
                            if len(magic_examples[num_str]) < 2:
                                col_letter = get_column_letter(col_idx)
                                magic_examples[num_str].append(
                                    f"{col_letter}{row_idx}"
                                )

            # Nur melden wenn dieselbe Magic Number oft vorkommt
            frequent = {v: c for v, c in magic_values.items() if c >= 3}
            if frequent:
                examples = ", ".join(
                    f"{v} ({c}x, z.B. {', '.join(magic_examples[v])})"
                    for v, c in sorted(frequent.items(), key=lambda x: -x[1])[:5]
                )
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=self.rule_name,
                    category=Category.FORMULA,
                    severity=Severity.WARNING,
                    message=(
                        f"Implizites Wissen in Formeln: {len(frequent)} "
                        f"hartcodierte Werte wiederholen sich in Formeln."
                    ),
                    sheet=ws.title,
                    detail=f"Werte: {examples}",
                    suggestion=(
                        "Tipp: Wiederkehrende Zahlen in Formeln (z.B. Steuersätze, "
                        "Faktoren, Limits) sollten als 'Benannter Bereich' oder in "
                        "einer Parametertabelle definiert werden. So wird beim Ändern "
                        "nur EINE Stelle angepasst statt vieler."
                    ),
                    score_penalty=min(8, len(frequent) * 2),
                ))

        return findings


class DataValidationRule(BaseRule):
    """Erkennt Datenvalidierungen mit hardcodierten Listen."""

    rule_id = "IMP-006"
    rule_name = "Hardcodierte Validierungslisten"
    needs_styles = True  # ws.data_validations ist im read_only-Modus nicht verfügbar

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            validations = ws.data_validations.dataValidation if ws.data_validations else []
            hardcoded_count = 0
            formula_based_count = 0

            for dv in validations:
                if dv.type == 'list' and dv.formula1:
                    formula = str(dv.formula1)
                    # Hardcodierte Liste: "A,B,C" vs. Bereichsreferenz: $A$1:$A$5
                    if ',' in formula and not re.search(r'[A-Z]+\d+', formula):
                        hardcoded_count += 1
                    else:
                        formula_based_count += 1

            if hardcoded_count > 3:
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=self.rule_name,
                    category=Category.STRUCTURE,
                    severity=Severity.WARNING,
                    message=(
                        f"{hardcoded_count} Dropdown-Listen mit eingetippten Werten – "
                        f"Änderungen müssen manuell in jeder Liste nachgezogen werden."
                    ),
                    sheet=ws.title,
                    suggestion=(
                        "Tipp: Dropdown-Werte in einer separaten Stammdaten-Tabelle "
                        "pflegen und per Bereichsreferenz einbinden. Dann muss bei "
                        "Änderungen nur EINE Stelle aktualisiert werden."
                    ),
                    score_penalty=min(5, hardcoded_count),
                ))

        return findings


class CrypticSheetNamesRule(BaseRule):
    """Erkennt generische oder kryptische Sheet-Namen."""

    rule_id = "IMP-007"
    rule_name = "Nicht-sprechende Sheet-Namen"

    GENERIC_PATTERNS = [
        re.compile(r'^(tabelle|sheet|blatt|mappe)\s*\d*$', re.IGNORECASE),
        re.compile(r'^(copy|kopie)\s+(of|von)\s+', re.IGNORECASE),
        re.compile(r'^\d+$'),  # Nur Zahlen
        re.compile(r'^(temp|tmp|test|neu|new|alt|old|backup)\s*\d*$', re.IGNORECASE),
        re.compile(r'\(\d+\)$'),  # "Sheet1 (2)"
    ]

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        cryptic = []

        for ws in workbook.worksheets:
            for pattern in self.GENERIC_PATTERNS:
                if pattern.search(ws.title):
                    cryptic.append(ws.title)
                    break

        if cryptic:
            findings.append(Finding(
                rule_id=self.rule_id,
                rule_name=self.rule_name,
                category=Category.STRUCTURE,
                severity=Severity.INFO,
                message=(
                    f"Nicht-sprechende Sheet-Namen: {', '.join(cryptic)}"
                ),
                suggestion=(
                    "Tipp: Aussagekräftige Sheet-Namen wie 'Umsatz_2024' oder "
                    "'Stammdaten_Kunden' helfen allen Beteiligten, sich schnell "
                    "zurechtzufinden. 'Tabelle1' sagt niemandem etwas."
                ),
                score_penalty=min(3, len(cryptic)),
            ))

        return findings


class CommentsWithLogicRule(BaseRule):
    """Erkennt Kommentare/Notizen, die Geschäftslogik enthalten könnten."""

    rule_id = "IMP-008"
    rule_name = "Geschäftslogik in Kommentaren"
    needs_styles = True  # cell.comment ist im read_only-Modus nicht verfügbar
    scales_with_cells = True

    LOGIC_KEYWORDS = {
        "achtung", "wichtig", "nicht ändern", "nicht löschen",
        "muss", "immer", "niemals", "regel", "formel",
        "wenn", "dann", "sonst", "berechnung", "prozent",
        "attention", "important", "do not", "must", "always",
        "never", "rule", "formula", "calculation",
    }

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            if ws.max_row is None:
                continue
            max_row = min(ws.max_row, 5000)
            max_col = min(ws.max_column or 1, 200)

            logic_comments = []

            for row_idx, row in iter_window_rows(ws, max_row, max_col):
                for col_pos, cell in enumerate(row):
                    col_idx = col_pos + 1
                    comment_text = ""
                    if getattr(cell, "comment", None):
                        comment_text = str(cell.comment.text or "").lower()

                    if comment_text:
                        has_logic = any(
                            kw in comment_text for kw in self.LOGIC_KEYWORDS
                        )
                        if has_logic:
                            col_letter = get_column_letter(col_idx)
                            logic_comments.append(
                                f"{col_letter}{row_idx}"
                            )

            if logic_comments:
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=self.rule_name,
                    category=Category.STRUCTURE,
                    severity=Severity.WARNING,
                    message=(
                        f"Geschäftslogik in Zell-Kommentaren entdeckt: "
                        f"{len(logic_comments)} Kommentare enthalten "
                        f"Hinweise wie 'Achtung', 'Nicht ändern' etc."
                    ),
                    sheet=ws.title,
                    detail=f"Zellen: {', '.join(logic_comments[:10])}",
                    suggestion=(
                        "Hinweis: Wenn kritische Regeln nur in Kommentaren stehen, "
                        "gehen sie leicht verloren oder werden übersehen. "
                        "Besser: Geschäftsregeln als Datenvalidierung, bedingte "
                        "Formatierung oder in einem Dokumentations-Sheet festhalten."
                    ),
                    score_penalty=min(5, len(logic_comments)),
                ))

        return findings


class CustomNumberFormatsRule(BaseRule):
    """Erkennt irreführende Zahlenformate (z.B. Zahlen die anders aussehen als sie sind)."""

    rule_id = "IMP-009"
    rule_name = "Irreführende Zahlenformate"
    needs_styles = True  # cell.number_format ist im read_only-Modus nicht verfügbar
    scales_with_cells = True
    supports_sampling = True

    SUSPICIOUS_FORMATS = [
        (r'0+\.0+.*["\']', "Zahl wird mit Text-Suffix angezeigt"),
        (r'#.*"', "Zahl wird mit Text-Zusatz angezeigt"),
        (r'0{4,}', "Übermäßig viele Dezimal-/Vorkommastellen"),
        (r'\[Red\]|\[Color', "Farbcodierung in Zahlenformat eingebettet"),
    ]

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            if ws.max_row is None or ws.max_row < 2:
                continue
            max_row = min(ws.max_row, 3000)
            max_col = min(ws.max_column or 1, 200)

            suspicious_formats: Counter = Counter()

            for _row_idx, row in iter_window_rows(ws, max_row, max_col):
                for cell in row:
                    if cell.number_format and cell.number_format != 'General':
                        fmt = cell.number_format
                        for pattern, desc in self.SUSPICIOUS_FORMATS:
                            if re.search(pattern, fmt, re.IGNORECASE):
                                suspicious_formats[f"{fmt} ({desc})"] += 1

            if suspicious_formats:
                format_summary = ", ".join(
                    f"'{fmt}' ({cnt}x)"
                    for fmt, cnt in suspicious_formats.most_common(5)
                )
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=self.rule_name,
                    category=Category.STRUCTURE,
                    severity=Severity.INFO,
                    message=(
                        f"Spezielle Zahlenformate entdeckt, die den angezeigten "
                        f"Wert vom tatsächlichen Wert abweichen lassen können."
                    ),
                    sheet=ws.title,
                    detail=f"Formate: {format_summary}",
                    suggestion=(
                        "Zur Info: Zahlenformate, die Text einblenden oder Werte "
                        "optisch verändern, können zu Missverständnissen führen "
                        "(z.B. '1' wird angezeigt, aber '1000' ist gespeichert). "
                        "Bitte die kritischen Zellen prüfen."
                    ),
                    score_penalty=2,
                ))

        return findings


class ProtectedAreasRule(BaseRule):
    """Erkennt Blattschutz – oft ein Zeichen für undokumentierte Regeln."""

    rule_id = "IMP-010"
    rule_name = "Blattschutz & gesperrte Bereiche"
    needs_styles = True  # ws.protection ist im read_only-Modus nicht verfügbar

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        protected_sheets = []

        for ws in workbook.worksheets:
            if ws.protection.sheet:
                protected_sheets.append(ws.title)

        if protected_sheets:
            findings.append(Finding(
                rule_id=self.rule_id,
                rule_name=self.rule_name,
                category=Category.STRUCTURE,
                severity=Severity.INFO,
                message=(
                    f"{len(protected_sheets)} Sheet(s) sind geschützt: "
                    f"{', '.join(protected_sheets[:5])}"
                ),
                suggestion=(
                    "Zur Info: Blattschutz zeigt, dass bestimmte Bereiche kritisch "
                    "sind. Gut so! Aber: Dokumentieren, WARUM und WAS geschützt ist. "
                    "Wenn der Ersteller das Unternehmen verlässt, kann der Blattschutz "
                    "zum Problem werden."
                ),
                score_penalty=1,
            ))

        return findings
