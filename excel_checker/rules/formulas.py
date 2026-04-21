"""Formel- und Bezugs-Regeln."""

from __future__ import annotations

import re
from collections import defaultdict
from typing import List

import openpyxl
from openpyxl.utils import get_column_letter

from excel_checker.i18n import _
from excel_checker.models import Category, Finding, Severity
from excel_checker.rules.base import BaseRule

# Volatile Funktionen (deutsch + englisch)
VOLATILE_FUNCTIONS = {
    "HEUTE", "TODAY", "JETZT", "NOW",
    "BEREICH.VERSCHIEBEN", "OFFSET",
    "INDIREKT", "INDIRECT",
    "ZUFALLSZAHL", "RAND", "ZUFALLSBEREICH", "RANDBETWEEN",
    "INFO", "ZELLE", "CELL",
}

# SVERWEIS/VLOOKUP und verwandte Lookup-Funktionen
LOOKUP_FUNCTIONS = {
    "SVERWEIS", "VLOOKUP", "WVERWEIS", "HLOOKUP",
    "VERWEIS", "LOOKUP", "INDEX", "VERGLEICH", "MATCH",
    "XVERWEIS", "XLOOKUP",
}

# Pattern für Zellbezüge in Formeln
CELL_REF_PATTERN = re.compile(
    r'(\$?)([A-Z]{1,3})(\$?)(\d{1,7})', re.IGNORECASE
)

# Pattern für Sheet-Referenzen
SHEET_REF_PATTERN = re.compile(
    r"(?:'([^']+)'|([A-Za-z0-9_äöüÄÖÜß]+))!", re.IGNORECASE
)

# Pattern für externe Datei-Referenzen
EXTERNAL_REF_PATTERN = re.compile(
    r"\[([^\]]+)\]", re.IGNORECASE
)

# Pattern für Funktionsaufrufe
FUNCTION_PATTERN = re.compile(
    r'([A-ZÄÖÜa-zäöü][A-ZÄÖÜa-zäöü0-9_.]+)\s*\(', re.IGNORECASE
)

# Muster für hartcodierte Zahlen in Formeln ("Zauberzahlen")
MAGIC_NUMBER_PATTERN = re.compile(
    r'(?<![A-Z$])(?<!\d)(\d+\.?\d*)(?![A-Z\d:$])', re.IGNORECASE
)


class AbsoluteVsRelativeRule(BaseRule):
    """Prüft das Verhältnis von absoluten ($) zu relativen Bezügen."""

    rule_id = "FRM-001"
    rule_name = _("FRM-001.name")
    scales_with_cells = True
    supports_sampling = True

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            if ws.max_row is None or ws.max_row < 2:
                continue
            max_row = min(ws.max_row, 5000)
            max_col = min(ws.max_column or 1, 200)

            absolute_count = 0
            relative_count = 0
            mixed_count = 0
            total_formulas = 0

            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.data_type != 'f' or not isinstance(cell.value, str):
                        continue
                    formula = cell.value
                    total_formulas += 1

                    refs = CELL_REF_PATTERN.findall(formula)
                    for col_dollar, col_let, row_dollar, row_num in refs:
                        if col_dollar and row_dollar:
                            absolute_count += 1
                        elif col_dollar or row_dollar:
                            mixed_count += 1
                        else:
                            relative_count += 1

            total_refs = absolute_count + relative_count + mixed_count
            if total_refs < 10:
                continue

            abs_pct = absolute_count / total_refs
            if abs_pct > 0.6:
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("FRM-001.name"),
                    category=Category.FORMULA,
                    severity=Severity.WARNING,
                    message=_("FRM-001.msg", pct=f"{abs_pct:.0%}", total=total_refs),
                    sheet=ws.title,
                    detail=_("FRM-001.detail",
                             absolute=absolute_count,
                             relative=relative_count,
                             mixed=mixed_count),
                    suggestion=_("FRM-001.tip"),
                    score_penalty=5,
                ))

        return findings


class VolatileFunctionsRule(BaseRule):
    """Erkennt volatile Funktionen, die Excel zur permanenten Neuberechnung zwingen."""

    rule_id = "FRM-002"
    rule_name = _("FRM-002.name")
    scales_with_cells = True
    supports_sampling = True

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            if ws.max_row is None or ws.max_row < 2:
                continue
            max_row = min(ws.max_row, 5000)
            max_col = min(ws.max_column or 1, 200)

            volatile_cells = defaultdict(list)

            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.data_type != 'f' or not isinstance(cell.value, str):
                        continue
                    functions_used = FUNCTION_PATTERN.findall(cell.value)
                    for func in functions_used:
                        if func.upper() in VOLATILE_FUNCTIONS:
                            col_letter = get_column_letter(col_idx)
                            volatile_cells[func.upper()].append(
                                f"{col_letter}{row_idx}"
                            )

            for func_name, cells in volatile_cells.items():
                cell_str = ', '.join(cells[:10]) + ('...' if len(cells) > 10 else '')
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("FRM-002.name"),
                    category=Category.FORMULA,
                    severity=Severity.WARNING if len(cells) < 50 else Severity.ERROR,
                    message=_("FRM-002.msg", func=func_name, count=len(cells)),
                    sheet=ws.title,
                    detail=_("FRM-002.detail", cells=cell_str),
                    suggestion=_("FRM-002.tip", func=func_name),
                    score_penalty=min(15, len(cells) // 5 + 3),
                ))

        return findings


class VlookupChainsRule(BaseRule):
    """Erkennt exzessive SVERWEIS/VLOOKUP-Nutzung als Datenbankersatz."""

    rule_id = "FRM-003"
    rule_name = _("FRM-003.name")
    scales_with_cells = True
    supports_sampling = True

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            if ws.max_row is None or ws.max_row < 2:
                continue
            max_row = min(ws.max_row, 5000)
            max_col = min(ws.max_column or 1, 200)

            lookup_count = 0
            lookup_functions_used = defaultdict(int)

            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.data_type != 'f' or not isinstance(cell.value, str):
                        continue
                    functions_used = FUNCTION_PATTERN.findall(cell.value)
                    for func in functions_used:
                        if func.upper() in LOOKUP_FUNCTIONS:
                            lookup_count += 1
                            lookup_functions_used[func.upper()] += 1

            if lookup_count > 50:
                func_summary = ", ".join(
                    f"{f}: {c}x" for f, c in
                    sorted(lookup_functions_used.items(), key=lambda x: -x[1])
                )
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("FRM-003.name"),
                    category=Category.FORMULA,
                    severity=Severity.WARNING if lookup_count < 500 else Severity.ERROR,
                    message=_("FRM-003.msg", count=lookup_count),
                    sheet=ws.title,
                    detail=_("FRM-003.detail", summary=func_summary),
                    suggestion=_("FRM-003.tip"),
                    score_penalty=min(20, lookup_count // 25),
                ))

        return findings


class CircularReferenceHintRule(BaseRule):
    """Sucht nach Zirkelverweisen (Formeln, die sich selbst referenzieren)."""

    rule_id = "FRM-004"
    rule_name = _("FRM-004.name")
    scales_with_cells = True

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            if ws.max_row is None or ws.max_row < 2:
                continue
            max_row = min(ws.max_row, 3000)
            max_col = min(ws.max_column or 1, 200)

            suspect_cells = []
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.data_type != 'f' or not isinstance(cell.value, str):
                        continue
                    col_letter = get_column_letter(col_idx)
                    own_ref = f"{col_letter}{row_idx}"
                    # Einfache Prüfung: Referenziert die Formel ihre eigene Zelle?
                    if own_ref in cell.value.upper().replace("$", ""):
                        suspect_cells.append(own_ref)

            if suspect_cells:
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("FRM-004.name"),
                    category=Category.FORMULA,
                    severity=Severity.ERROR,
                    message=_("FRM-004.msg", count=len(suspect_cells)),
                    sheet=ws.title,
                    detail=_("FRM-004.detail", cells=', '.join(suspect_cells[:10])),
                    suggestion=_("FRM-004.tip"),
                    score_penalty=min(15, len(suspect_cells) * 3),
                ))

        return findings


class CrossSheetReferenceRule(BaseRule):
    """Analysiert Abhängigkeiten zwischen Sheets und zu externen Dateien."""

    rule_id = "FRM-005"
    rule_name = _("FRM-005.name")
    scales_with_cells = True
    supports_sampling = True

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            if ws.max_row is None or ws.max_row < 2:
                continue
            max_row = min(ws.max_row, 5000)
            max_col = min(ws.max_column or 1, 200)

            sheet_refs = defaultdict(int)
            external_refs = defaultdict(int)

            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.data_type != 'f' or not isinstance(cell.value, str):
                        continue
                    formula = cell.value

                    # Externe Dateien
                    for ext_match in EXTERNAL_REF_PATTERN.finditer(formula):
                        external_refs[ext_match.group(1)] += 1

                    # Andere Sheets
                    for sheet_match in SHEET_REF_PATTERN.finditer(formula):
                        ref_sheet = sheet_match.group(1) or sheet_match.group(2)
                        if ref_sheet != ws.title:
                            sheet_refs[ref_sheet] += 1

            if external_refs:
                files = ", ".join(
                    f"'{f}' ({c}x)" for f, c in
                    sorted(external_refs.items(), key=lambda x: -x[1])[:5]
                )
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("FRM-005.name"),
                    category=Category.FORMULA,
                    severity=Severity.ERROR,
                    message=_("FRM-005.msg.external", count=sum(external_refs.values())),
                    sheet=ws.title,
                    detail=_("FRM-005.detail.external", files=files),
                    suggestion=_("FRM-005.tip.external"),
                    score_penalty=min(15, sum(external_refs.values()) // 5 + 5),
                ))

            if len(sheet_refs) > 5:
                sheets = ", ".join(
                    f"'{s}' ({c}x)" for s, c in
                    sorted(sheet_refs.items(), key=lambda x: -x[1])[:5]
                )
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("FRM-005.name"),
                    category=Category.FORMULA,
                    severity=Severity.INFO,
                    message=_("FRM-005.msg.cross_sheet", count=len(sheet_refs)),
                    sheet=ws.title,
                    detail=_("FRM-005.detail.cross_sheet", sheets=sheets),
                    suggestion=_("FRM-005.tip.cross_sheet"),
                    score_penalty=2,
                ))

        return findings
