"""Struktur- und Normalform-Regeln (inkl. Identifier-Konsistenz)."""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from typing import List, Optional

import openpyxl
from openpyxl.utils import get_column_letter

from excel_checker.i18n import _
from excel_checker.models import Category, Finding, Severity
from excel_checker.rules._sampling import iter_window_rows
from excel_checker.rules.base import BaseRule


def _scan_columns(ws, max_row: int, max_col: int, text_only: bool = False):
    """Liest Kopfzeile und Spaltenwerte in einem einzigen Vorwaertslauf.

    Ersetzt das frueher spaltenweise ``ws.cell(row, col)``-Muster, das im
    ``read_only``-Modus nicht funktioniert.

    ``text_only=False`` -> je Spalte die Liste der nicht-leeren Werte.
    ``text_only=True``  -> je Spalte nur String-Werte als ``(row_idx, wert)``.
    """
    headers: List[Optional[object]] = [None] * max_col
    values_per_col: List[list] = [[] for _c in range(max_col)]
    for row_idx, row in iter_window_rows(ws, max_row, max_col):
        if row_idx == 1:
            for col_pos, cell in enumerate(row):
                headers[col_pos] = cell.value
            continue
        for col_pos, cell in enumerate(row):
            value = cell.value
            if value is None:
                continue
            if text_only:
                if isinstance(value, str):
                    values_per_col[col_pos].append((row_idx, value.strip()))
            else:
                values_per_col[col_pos].append(value)
    return headers, values_per_col


class MergedCellsRule(BaseRule):
    """Erkennt verbundene Zellen – der Feind maschineller Lesbarkeit."""

    rule_id = "STR-001"
    rule_name = _("STR-001.name")
    needs_styles = True  # ws.merged_cells ist im read_only-Modus nicht verfügbar

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            merged = list(ws.merged_cells.ranges)
            if merged:
                examples = [str(r) for r in merged[:5]]
                extra = f" (und {len(merged) - 5} weitere)" if len(merged) > 5 else ""
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("STR-001.name"),
                    category=Category.STRUCTURE,
                    severity=Severity.ERROR if len(merged) > 10 else Severity.WARNING,
                    message=_("STR-001.msg", count=len(merged)),
                    sheet=ws.title,
                    detail=_("STR-001.detail", ranges=f"{', '.join(examples)}{extra}"),
                    suggestion=_("STR-001.tip"),
                    score_penalty=min(20, len(merged) * 2),
                ))
        return findings


class DataTypeHomogeneityRule(BaseRule):
    """Prüft ob Spalten einheitliche Datentypen enthalten (1. Normalform)."""

    rule_id = "STR-002"
    rule_name = _("STR-002.name")
    scales_with_cells = True
    supports_sampling = True

    def check(
        self,
        workbook: openpyxl.Workbook,
        file_path: str,
        progress_callback=None,
        sample_mode=None,
    ) -> List[Finding]:
        findings = []
        # Bei aktiver Stichprobe: engere Schranken, sonst die bisherigen Limits.
        if sample_mode is not None:
            row_cap = min(sample_mode.max_rows_per_sheet, 3000)
            col_cap = min(sample_mode.max_cols_per_sheet, 100)
            sample_note = sample_mode.disclosure_de()
        else:
            row_cap = 3000
            col_cap = 100
            sample_note = None

        for ws in workbook.worksheets:
            if ws.max_row is None or ws.max_row < 3:
                continue
            max_row = min(ws.max_row, row_cap)
            max_col = min(ws.max_column or 1, col_cap)
            # Ein Vorwaertslauf ab Zeile 2, Spaltenstatistik nebenher.
            counts_per_col: List[Counter] = [Counter() for _c in range(max_col)]
            non_empty_per_col = [0] * max_col
            for _row_idx, row in iter_window_rows(ws, max_row, max_col, min_row=2):
                for col_pos, cell in enumerate(row):
                    value = cell.value
                    if value is None:
                        continue
                    non_empty_per_col[col_pos] += 1
                    if isinstance(value, (int, float)):
                        counts_per_col[col_pos]["Zahl"] += 1
                    elif isinstance(value, str):
                        counts_per_col[col_pos]["Text"] += 1
                    elif isinstance(value, bool):
                        counts_per_col[col_pos]["Boolean"] += 1
                    else:
                        counts_per_col[col_pos]["Datum/Andere"] += 1

            for col_idx in range(1, max_col + 1):
                type_counts = counts_per_col[col_idx - 1]
                non_empty = non_empty_per_col[col_idx - 1]

                if non_empty < 5:
                    continue

                # Prüfe ob ein Typ < 80% dominant ist
                dominant_type = type_counts.most_common(1)[0] if type_counts else None
                if dominant_type and len(type_counts) > 1:
                    dominant_pct = dominant_type[1] / non_empty
                    if dominant_pct < 0.95:
                        col_letter = get_column_letter(col_idx)
                        minority_types = {k: v for k, v in type_counts.items()
                                          if k != dominant_type[0]}
                        findings.append(Finding(
                            rule_id=self.rule_id,
                            rule_name=_("STR-002.name"),
                            category=Category.STRUCTURE,
                            severity=Severity.WARNING,
                            message=_("STR-002.msg",
                                      col=col_letter,
                                      dominant=dominant_type[0],
                                      pct=f"{dominant_pct:.0%}",
                                      minority=minority_types),
                            sheet=ws.title,
                            suggestion=_("STR-002.tip", col=col_letter),
                            score_penalty=3,
                            sample_note=sample_note,
                        ))
        return findings


class HeaderDetectionRule(BaseRule):
    """Prüft ob eine saubere einzelne Kopfzeile existiert."""

    rule_id = "STR-003"
    rule_name = _("STR-003.name")

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            if ws.max_row is None or ws.max_row < 2:
                continue

            # Die ersten 5 Zeilen einmal lesen — beide Prüfungen unten
            # arbeiten auf demselben Fenster.
            max_col = min(ws.max_column or 1, 99)
            head_rows = {
                row_idx: [c.value for c in row]
                for row_idx, row in iter_window_rows(
                    ws, min(5, ws.max_row), max_col
                )
            }

            header_candidates = []
            for row_idx, row_values in head_rows.items():
                non_empty = sum(1 for v in row_values if v is not None)
                all_text = all(isinstance(v, str) for v in row_values if v is not None)
                if non_empty > 0 and all_text:
                    header_candidates.append(row_idx)

            if not header_candidates:
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("STR-003.name"),
                    category=Category.STRUCTURE,
                    severity=Severity.WARNING,
                    message=_("STR-003.msg.none"),
                    sheet=ws.title,
                    suggestion=_("STR-003.tip.none"),
                    score_penalty=5,
                ))
            elif header_candidates[0] != 1:
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("STR-003.name"),
                    category=Category.STRUCTURE,
                    severity=Severity.WARNING,
                    message=_("STR-003.msg.late", row=header_candidates[0]),
                    sheet=ws.title,
                    suggestion=_("STR-003.tip.late"),
                    score_penalty=3,
                ))

            # Prüfe auf doppelte Header-Namen
            if header_candidates:
                hr = header_candidates[0]
                names = [
                    str(val).strip()
                    for val in head_rows.get(hr, [])
                    if val is not None
                ]
                dupes = [n for n, cnt in Counter(names).items() if cnt > 1]
                if dupes:
                    findings.append(Finding(
                        rule_id=self.rule_id,
                        rule_name=_("STR-003.name"),
                        category=Category.STRUCTURE,
                        severity=Severity.ERROR,
                        message=_("STR-003.msg.dupes", dupes=', '.join(dupes)),
                        sheet=ws.title,
                        suggestion=_("STR-003.tip.dupes"),
                        score_penalty=8,
                    ))
        return findings


class EmptyRowColSeparatorRule(BaseRule):
    """Erkennt leere Zeilen/Spalten als Trenner im Datenbereich."""

    rule_id = "STR-004"
    rule_name = _("STR-004.name")
    scales_with_cells = True

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            if ws.max_row is None or ws.max_row < 5:
                continue
            max_row = min(ws.max_row, 5000)
            max_col = min(ws.max_column or 1, 200)

            # Leere Zeilen im Datenbereich
            empty_rows = []
            data_started = False
            data_ended = False
            for row_idx, row in iter_window_rows(ws, max_row, max_col):
                is_empty = all(cell.value is None for cell in row)
                if not is_empty:
                    if data_ended:
                        # Daten nach einer Lücke = fragmentiert
                        empty_rows.append(row_idx)
                    data_started = True
                    data_ended = False
                elif data_started:
                    data_ended = True
                    empty_rows.append(row_idx)

            if len(empty_rows) > 2:
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("STR-004.name"),
                    category=Category.STRUCTURE,
                    severity=Severity.WARNING,
                    message=_("STR-004.msg", count=len(empty_rows)),
                    sheet=ws.title,
                    suggestion=_("STR-004.tip"),
                    score_penalty=min(10, len(empty_rows)),
                ))
        return findings


class IdentifierConsistencyRule(BaseRule):
    """Prüft ob Identifier-Spalten konsistente Formate haben (z.B. DEM-001 vs DEM-2)."""

    rule_id = "STR-005"
    rule_name = _( "STR-005.name")
    scales_with_cells = True
    supports_sampling = True

    # Pattern das typische IDs erkennt: PREFIX-NNN, PREFIX_NNN, PREFIX.NNN
    ID_PATTERN = re.compile(r'^([A-Za-zÄÖÜäöü]+)([-_./])(\d+)(.*)$')

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws_idx, ws in enumerate(workbook.worksheets):
            if ws.max_row is None or ws.max_row < 3:
                continue
            max_row = min(ws.max_row, 3000)
            max_col = min(ws.max_column or 1, 100)

            # Textwerte aller Spalten in einem Lauf sammeln.
            values_per_col: List[list] = [[] for _c in range(max_col)]
            for row_idx, row in iter_window_rows(ws, max_row, max_col, min_row=2):
                for col_pos, cell in enumerate(row):
                    if isinstance(cell.value, str):
                        values_per_col[col_pos].append((row_idx, cell.value.strip()))
                # Fortschritt melden (alle 500 Zeilen)
                if progress_callback and row_idx % 500 == 0:
                    progress_callback({
                        "rule": self.rule_id,
                        "sheet": ws.title,
                        "row": row_idx,
                        "status": "scanning"
                    })

            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                values = values_per_col[col_idx - 1]

                if len(values) < 3:
                    continue

                # Prüfe ob die Spalte Identifier enthält
                id_matches = []
                for row_idx, val in values:
                    m = self.ID_PATTERN.match(val)
                    if m:
                        id_matches.append((row_idx, val, m))

                # Nur wenn > 50% der Werte wie IDs aussehen
                if len(id_matches) < len(values) * 0.5:
                    continue

                # Analysiere Konsistenz
                issues = self._check_consistency(id_matches, col_letter)
                findings.extend(
                    Finding(
                        rule_id=self.rule_id,
                        rule_name=self.rule_name,
                        category=Category.STRUCTURE,
                        severity=sev,
                        message=msg,
                        sheet=ws.title,
                        detail=detail,
                        suggestion=suggestion,
                        score_penalty=penalty,
                    )
                    for sev, msg, detail, suggestion, penalty in issues
                )

                # Duplikate-Check
                dupes = self._check_duplicates(id_matches, col_letter)
                findings.extend(
                    Finding(
                        rule_id=self.rule_id,
                        rule_name=_("STR-005.name"),
                        category=Category.STRUCTURE,
                        severity=Severity.ERROR,
                        message=msg,
                        sheet=ws.title,
                        detail=detail,
                        suggestion=_("STR-005.tip.dupes"),
                        score_penalty=8,
                    )
                    for msg, detail in dupes
                )

        return findings

    def _check_consistency(self, id_matches, col_letter):
        issues = []

        # Gruppiere nach Prefix+Separator
        prefix_groups: dict[str, list] = defaultdict(list)
        for row_idx, val, m in id_matches:
            key = (m.group(1).upper(), m.group(2))
            prefix_groups[key].append((row_idx, val, m.group(3)))

        for (prefix, sep), entries in prefix_groups.items():
            # Prüfe Ziffernbreite (z.B. 001 vs 2 vs 0003)
            digit_lengths = set(len(num_str) for _, _, num_str in entries)
            if len(digit_lengths) > 1:
                # Wichtig: KEIN `_` als Iterationsvariable — `_` ist oben als
                # i18n-Funktion importiert und würde hier überschrieben, sodass
                # nachfolgende _("STR-005...")-Calls mit TypeError crashen.
                examples = [val for _row_idx, val, _num in entries[:5]]
                issues.append((
                    Severity.ERROR,
                    _("STR-005.msg.inconsistent", col=col_letter, prefix=prefix, sep=sep),
                    _("STR-005.detail.inconsistent",
                      widths=str(sorted(digit_lengths)),
                      examples=', '.join(examples)),
                    _("STR-005.tip.inconsistent", prefix=prefix, sep=sep),
                    10,
                ))

            # Prüfe auf Lücken in der Nummerierung
            numbers = sorted(int(num_str) for _, _, num_str in entries)
            if len(numbers) > 3:
                expected = set(range(numbers[0], numbers[-1] + 1))
                missing = expected - set(numbers)
                if missing and len(missing) <= 20:
                    missing_str = str(sorted(missing)[:10]) + ('...' if len(missing) > 10 else '')
                    issues.append((
                        Severity.WARNING,
                        _("STR-005.msg.gaps", col=col_letter, prefix=prefix, sep=sep),
                        _("STR-005.detail.gaps", missing=missing_str),
                        _("STR-005.tip.gaps"),
                        3,
                    ))

        # Prüfe ob verschiedene Separatoren für gleichen Prefix verwendet werden
        prefixes_only = defaultdict(set)
        for (prefix, sep) in prefix_groups:
            prefixes_only[prefix].add(sep)
        for prefix, seps in prefixes_only.items():
            if len(seps) > 1:
                issues.append((
                    Severity.ERROR,
                    _("STR-005.msg.separators", col=col_letter, prefix=prefix, seps=seps),
                    None,
                    _("STR-005.tip.separators", prefix=prefix),
                    8,
                ))

        return issues

    def _check_duplicates(self, id_matches, col_letter):
        issues = []
        seen = Counter(val for _, val, _ in id_matches)
        dupes = {val: cnt for val, cnt in seen.items() if cnt > 1}
        if dupes:
            examples = list(dupes.items())[:5]
            issues.append((
                _("STR-005.msg.dupes", col=col_letter),
                _("STR-005.detail.dupes", dupes=', '.join(f'{v} ({c}x)' for v, c in examples)),
            ))
        return issues


class MissingUniqueKeyRule(BaseRule):
    """Erkennt Datentabellen ohne eindeutige Schlüsselspalte (Primärschlüssel)."""

    rule_id = "STR-006"
    rule_name = _("STR-006.name")
    scales_with_cells = True

    # Header-Begriffe die NICHT als ID-Spalten zählen
    NON_ID_HEADERS = {
        "bemerkung", "kommentar", "notiz", "beschreibung", "anmerkung",
        "comment", "note", "description", "remark",
    }

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            if ws.max_row is None or ws.max_row < 12:
                continue
            max_row = min(ws.max_row, 3000)
            max_col = min(ws.max_column or 1, 100)

            # Nur Sheets prüfen die wie Datentabellen aussehen
            data_rows = max_row - 1  # minus Header
            if data_rows < 10 or max_col < 2:
                continue

            headers, values_per_col = _scan_columns(ws, max_row, max_col)

            has_unique_col = False
            for col_idx in range(1, max_col + 1):
                header = headers[col_idx - 1]
                if isinstance(header, str) and header.strip().lower() in self.NON_ID_HEADERS:
                    continue

                values = values_per_col[col_idx - 1]

                if len(values) < data_rows * 0.8:
                    continue  # Zu viele Lücken

                if len(set(values)) == len(values):
                    has_unique_col = True
                    break

            if not has_unique_col:
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("STR-006.name"),
                    category=Category.STRUCTURE,
                    severity=Severity.CRITICAL,
                    message=_("STR-006.msg"),
                    sheet=ws.title,
                    detail=_("STR-006.detail", rows=data_rows, cols=max_col),
                    suggestion=_("STR-006.tip"),
                    score_penalty=15,
                ))
        return findings


class TextBasedIdRule(BaseRule):
    """Erkennt ID-Spalten mit Freitext-Werten, die nicht sortierbar sind."""

    rule_id = "STR-007"
    rule_name = _("STR-007.name")
    scales_with_cells = True
    supports_sampling = True

    # Header-Begriffe die auf ID-Spalten hindeuten
    ID_HEADER_HINTS = {
        "id", "nr", "nr.", "nummer", "number", "key", "schlüssel",
        "code", "kennung", "index", "#", "lfd", "lfd.",
        "bezeichnung", "name", "kürzel",
    }

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            if ws.max_row is None or ws.max_row < 10:
                continue
            max_row = min(ws.max_row, 3000)
            max_col = min(ws.max_column or 1, 100)

            headers, text_values = _scan_columns(ws, max_row, max_col, text_only=True)

            for col_idx in range(1, max_col + 1):
                header = headers[col_idx - 1]
                if not isinstance(header, str):
                    continue

                header_lower = header.strip().lower()
                is_id_col = any(hint in header_lower for hint in self.ID_HEADER_HINTS)
                if not is_id_col:
                    continue

                values = text_values[col_idx - 1]

                if len(values) < 5:
                    continue

                # Prüfe ob Werte Freitext sind (lang, mit Leerzeichen)
                long_text_count = 0
                space_count = 0
                # KEIN `_` als Iterationsvariable — würde die i18n-Funktion
                # überschreiben. Siehe IdentifierConsistencyRule._check_consistency.
                for _row_idx, val in values:
                    if len(val) > 30:
                        long_text_count += 1
                    if val.count(" ") >= 2:
                        space_count += 1

                freetext_ratio = max(long_text_count, space_count) / len(values)
                if freetext_ratio > 0.3 and len(values) >= 5:
                    col_letter = get_column_letter(col_idx)
                    examples = [val for _, val in values[:3]]
                    findings.append(Finding(
                        rule_id=self.rule_id,
                        rule_name=_("STR-007.name"),
                        category=Category.STRUCTURE,
                        severity=Severity.ERROR,
                        message=_("STR-007.msg", col=col_letter, header=header),
                        sheet=ws.title,
                        detail=_("STR-007.detail",
                                 examples=', '.join(repr(e) for e in examples),
                                 ratio=f"{freetext_ratio:.0%}"),
                        suggestion=_("STR-007.tip"),
                        score_penalty=12,
                    ))
        return findings
