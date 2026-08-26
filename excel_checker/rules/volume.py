"""Volumen- und Limit-Regeln."""

from __future__ import annotations

import os
from typing import List

import openpyxl
from openpyxl.utils import get_column_letter

from excel_checker.i18n import _
from excel_checker.models import Category, Finding, Severity
from excel_checker.rules._sampling import iter_window_rows
from excel_checker.rules.base import BaseRule


class RowCountRule(BaseRule):
    """Prüft ob Datenmengen das sinnvolle Excel-Limit überschreiten."""

    rule_id = "VOL-001"
    rule_name = _("VOL-001.name")

    THRESHOLDS = [
        (100_000, Severity.CRITICAL, "VOL-001.threshold.100k"),
        (50_000, Severity.ERROR, "VOL-001.threshold.50k"),
        (10_000, Severity.WARNING, "VOL-001.threshold.10k"),
    ]

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            if ws.max_row is None:
                continue

            for threshold, severity, label_key in self.THRESHOLDS:
                if ws.max_row > threshold:
                    findings.append(Finding(
                        rule_id=self.rule_id,
                        rule_name=_("VOL-001.name"),
                        category=Category.VOLUME,
                        severity=severity,
                        message=_("VOL-001.msg",
                                  rows=f"{ws.max_row:,}",
                                  label=_(label_key)),
                        sheet=ws.title,
                        detail=_("VOL-001.detail",
                                 rows=f"{ws.max_row:,}",
                                 cols=ws.max_column or 0),
                        suggestion=_("VOL-001.tip"),
                        score_penalty=15 if severity == Severity.CRITICAL else
                                      10 if severity == Severity.ERROR else 5,
                    ))
                    break  # Nur die höchste Stufe melden
        return findings


class FormulaDensityRule(BaseRule):
    """Prüft die Formeldichte – zu viele Formeln machen Excel instabil."""

    rule_id = "VOL-002"
    rule_name = _("VOL-002.name")

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        for ws in workbook.worksheets:
            if ws.max_row is None or ws.max_row < 5:
                continue
            max_row = min(ws.max_row, 3000)
            max_col = min(ws.max_column or 1, 100)

            formula_count = 0
            static_count = 0

            for _row_idx, row in iter_window_rows(ws, max_row, max_col):
                for cell in row:
                    if cell.value is None:
                        continue
                    if cell.data_type == 'f':
                        formula_count += 1
                    else:
                        static_count += 1

            total = formula_count + static_count
            if total < 100:
                continue

            formula_pct = formula_count / total if total > 0 else 0

            if formula_count > 10000:
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("VOL-002.name"),
                    category=Category.VOLUME,
                    severity=Severity.ERROR,
                    message=_("VOL-002.msg.high",
                              count=f"{formula_count:,}",
                              pct=f"{formula_pct:.0%}"),
                    sheet=ws.title,
                    detail=_("VOL-002.detail",
                             formulas=f"{formula_count:,}",
                             statics=f"{static_count:,}"),
                    suggestion=_("VOL-002.tip.high"),
                    score_penalty=min(15, formula_count // 2000),
                ))
            elif formula_count > 2000:
                findings.append(Finding(
                    rule_id=self.rule_id,
                    rule_name=_("VOL-002.name"),
                    category=Category.VOLUME,
                    severity=Severity.WARNING,
                    message=_("VOL-002.msg.medium", count=f"{formula_count:,}"),
                    sheet=ws.title,
                    suggestion=_("VOL-002.tip.medium"),
                    score_penalty=3,
                ))

        return findings


class SheetCountRule(BaseRule):
    """Prüft ob zu viele Tabellenblätter vorhanden sind."""

    rule_id = "VOL-003"
    rule_name = _("VOL-003.name")

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        sheet_count = len(workbook.sheetnames)

        if sheet_count > 20:
            sheets_str = ', '.join(workbook.sheetnames[:10]) + ('...' if sheet_count > 10 else '')
            findings.append(Finding(
                rule_id=self.rule_id,
                rule_name=_("VOL-003.name"),
                category=Category.VOLUME,
                severity=Severity.WARNING if sheet_count < 50 else Severity.ERROR,
                message=_("VOL-003.msg", count=sheet_count),
                detail=_("VOL-003.detail", sheets=sheets_str),
                suggestion=_("VOL-003.tip"),
                score_penalty=min(10, sheet_count // 5),
            ))

        return findings


class FileSizeRule(BaseRule):
    """Prüft die Dateigröße."""

    rule_id = "VOL-004"
    rule_name = _("VOL-004.name")

    def check(self, workbook: openpyxl.Workbook, file_path: str, progress_callback=None) -> List[Finding]:
        findings = []
        # Der Engine reicht die Größe durch; bei einem Puffer gibt es keinen
        # Pfad, den os.path.getsize vermessen könnte.
        size_bytes = getattr(self, "file_size_bytes", None)
        if size_bytes is None:
            try:
                size_bytes = os.path.getsize(file_path)
            except (OSError, TypeError, ValueError):
                return findings

        size_mb = size_bytes / (1024 * 1024)

        if size_mb > 50:
            findings.append(Finding(
                rule_id=self.rule_id,
                rule_name=_("VOL-004.name"),
                category=Category.VOLUME,
                severity=Severity.CRITICAL,
                message=_("VOL-004.msg.critical", size=f"{size_mb:.1f}"),
                suggestion=_("VOL-004.tip.critical"),
                score_penalty=20,
            ))
        elif size_mb > 20:
            findings.append(Finding(
                rule_id=self.rule_id,
                rule_name=_("VOL-004.name"),
                category=Category.VOLUME,
                severity=Severity.WARNING,
                message=_("VOL-004.msg.warning", size=f"{size_mb:.1f}"),
                suggestion=_("VOL-004.tip.warning"),
                score_penalty=8,
            ))

        return findings
