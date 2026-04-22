"""Column-Profiling für Migration/AI-Readiness-Bewertung (Phase 2 A.2).

Baut pro Daten-Spalte ein Profil: Null-Ratio, Datentyp-Mix, Unique-Count.
Ergebnisse werden als Findings nur bei echten Problemen (>50% Null in
produktiver Spalte) emittiert. Die eigentliche Profil-Liste landet auf
``self.column_profiles`` und wird vom Engine nach Ausführung in
``report.column_profiles`` übertragen.
"""

from __future__ import annotations

from collections import Counter
from datetime import date, datetime, time
from typing import List, Optional

import openpyxl
from openpyxl.utils import get_column_letter

from excel_checker.models import Category, ColumnProfile, Finding, Severity
from excel_checker.rules._sampling import SampleMode, iter_sampled_rows
from excel_checker.rules.base import BaseRule


def _classify(value) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    if isinstance(value, (datetime, date, time)):
        return "date"
    if isinstance(value, str):
        return "str"
    return "other"


class ColumnProfileRule(BaseRule):
    """Profiliert Spalten für Migration/AI-Readiness.

    Emittiert:
    - Spalten-Profile auf self.column_profiles (Engine liest diese ab).
    - WARNING-Findings bei Spalten mit >50% Null-Ratio, die vermutlich nicht
      bewusst optional sind.
    """

    rule_id = "PRF-001"
    rule_name = "Spalten-Profiling für Migration"
    scales_with_cells = True
    # Profile brauchen eine vollständige Sicht pro Spalte, nicht nur eine
    # zufällige Stichprobe. Wenn Sampling aktiv wird, produziert der Engine
    # stattdessen keine Profile (Regel wird in Tier 3 ohnehin gefiltert, da
    # sie scales_with_cells=True + supports_sampling=False setzt).
    supports_sampling = False

    _max_rows_scan = 10_000
    _max_cols_scan = 200

    def __init__(self) -> None:
        self.column_profiles: list[ColumnProfile] = []

    def check(self, workbook: openpyxl.Workbook, file_path: str,
              progress_callback=None, sample_mode: Optional[SampleMode] = None) -> List[Finding]:
        self.column_profiles = []
        findings: List[Finding] = []
        for ws in workbook.worksheets:
            max_row = min(ws.max_row or 0, self._max_rows_scan)
            max_col = min(ws.max_column or 0, self._max_cols_scan)
            if max_row < 3 or max_col < 1:
                continue
            headers: list[Optional[str]] = [None] * max_col
            first_row = next(iter_sampled_rows(ws, 1, max_col, None), None)
            if first_row:
                _, cells = first_row
                for idx, cell in enumerate(cells):
                    val = cell.value
                    headers[idx] = str(val).strip() if val not in (None, "") else None

            type_per_col: list[Counter] = [Counter() for _ in range(max_col)]
            uniques_per_col: list[set] = [set() for _ in range(max_col)]
            totals: list[int] = [0] * max_col

            for row_idx, row in iter_sampled_rows(ws, max_row, max_col, None):
                if row_idx == 1:
                    continue
                for col_idx, cell in enumerate(row):
                    val = cell.value
                    totals[col_idx] += 1
                    type_per_col[col_idx][_classify(val)] += 1
                    if val is not None and not isinstance(val, bool):
                        uniques_per_col[col_idx].add(val)

            for col_idx in range(max_col):
                if totals[col_idx] == 0:
                    continue
                type_mix = dict(type_per_col[col_idx])
                nulls = type_mix.get("null", 0)
                non_null_types = {k: v for k, v in type_mix.items() if k != "null"}
                dominant = None
                if non_null_types:
                    total_non_null = sum(non_null_types.values())
                    top_type, top_count = max(non_null_types.items(), key=lambda kv: kv[1])
                    dominant = top_type if top_count / total_non_null >= 0.8 else "mixed"

                profile = ColumnProfile(
                    sheet=ws.title,
                    column_letter=get_column_letter(col_idx + 1),
                    header=headers[col_idx],
                    total=totals[col_idx],
                    nulls=nulls,
                    unique_count=len(uniques_per_col[col_idx]),
                    dominant_type=dominant,
                    type_mix=type_mix,
                )
                self.column_profiles.append(profile)

                # Soft-Finding bei extrem hoher Null-Quote in anscheinend
                # produktiven Spalten (mindestens 20 Einträge).
                if profile.total >= 20 and profile.null_ratio > 0.5:
                    col_letter = profile.column_letter
                    findings.append(Finding(
                        rule_id="PRF-002",
                        rule_name="Hohe Null-Quote in Spalte",
                        category=Category.STRUCTURE,
                        severity=Severity.INFO,
                        message=(
                            f"Spalte {col_letter} ist zu {profile.null_ratio:.0%} leer "
                            f"(Header: {profile.header or '—'})"
                        ),
                        sheet=ws.title,
                        detail=(
                            f"Nur {profile.total - profile.nulls} von {profile.total} Zeilen "
                            f"sind befüllt. Dominanter Typ: {profile.dominant_type or '—'}."
                        ),
                        suggestion=(
                            "Wenn die Spalte optional ist, Header als 'optional' kennzeichnen. "
                            "Wenn sie eigentlich gefüllt sein sollte, fehlende Werte nachtragen."
                        ),
                        score_penalty=1,
                    ))
        return findings
