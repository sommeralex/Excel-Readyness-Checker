"""Haupt-Engine: Lädt Excel, führt Regeln aus, generiert Bericht."""

from __future__ import annotations

import json
import os
import time
from typing import Callable, Generator, List, Optional, Tuple

import openpyxl

from excel_checker.i18n import _
from excel_checker.models import SheetStats, WorkbookReport
from excel_checker.recommendations import generate_recommendations
from excel_checker.rules import ALL_RULES, BaseRule

# Excel physische Limits – Dimensionen an diesen Werten = Phantom-Ausdehnung
EXCEL_MAX_ROW = 1_048_576  # 2^20
EXCEL_MAX_COL = 16_384     # 2^14
_EMPTY_RUN_EXIT = 200      # Nach N leeren Zeilen → Datenende

# Schwelle ab der wir max_row/max_col als "aufgebläht" betrachten und
# per Fuzzy-Scan die echte Datengrenze suchen statt blindem Hochrechnen.
_SUSPICIOUS_ROW_THRESHOLD = 5_000   # > 5000 gemeldete Zeilen → prüfen
_SUSPICIOUS_COL_THRESHOLD = 200     # > 200 gemeldete Spalten → prüfen


# ── Fuzzy-Scan: Schnelles Auffinden der echten Datengrenzen ──────────

def _row_has_data(ws, row_idx: int, max_col: int) -> bool:
    """Prüft ob eine Zeile in den ersten max_col Spalten Daten enthält."""
    for col_idx in range(1, min(max_col, 100) + 1):
        if ws.cell(row=row_idx, column=col_idx).value is not None:
            return True
    return False


def _find_last_data_row(ws, max_row: int, check_cols: int) -> int:
    """Findet die letzte befüllte Zeile per Fuzzy-Binärsuche.

    Strategie:
    1. Erste 100 Zeilen linear prüfen (Header + Datenanfang)
    2. Ab Zeile 100 in exponentiell wachsenden Schritten vorspringen
    3. Wenn leer: per Binärsuche die exakte Grenze zwischen letztem
       Sprung mit Daten und erstem Sprung ohne Daten finden
    4. Von der gefundenen Grenze noch 50 Zeilen linear zurücklaufen
       um vereinzelte leere Zeilen am Ende nicht mitzuzählen
    """
    check_cols = min(check_cols, 100)

    # Phase 1: Erste 100 Zeilen linear
    last_found = 0
    linear_end = min(100, max_row)
    for r in range(1, linear_end + 1):
        if _row_has_data(ws, r, check_cols):
            last_found = r

    if max_row <= linear_end:
        return last_found

    # Phase 2: Exponentielle Sprünge (200, 400, 800, 1600, ...)
    prev_pos = linear_end
    step = 200
    probe_pos = linear_end + step
    last_data_probe = last_found  # Letzte Probe-Position MIT Daten

    while probe_pos <= max_row:
        if _row_has_data(ws, probe_pos, check_cols):
            last_data_probe = probe_pos
            prev_pos = probe_pos
            step = min(step * 2, 10_000)  # Cap bei 10k Schritten
            probe_pos += step
        else:
            # Erste leere Probe gefunden → Binärsuche zwischen prev_pos und probe_pos
            break
    else:
        # Nie eine leere Probe gefunden → Daten gehen bis max_row
        # Trotzdem: letzten Bereich linear prüfen
        probe_pos = max_row

    # Phase 3: Binärsuche zwischen last_data_probe und probe_pos
    lo, hi = last_data_probe, min(probe_pos, max_row)
    while hi - lo > 50:
        mid = (lo + hi) // 2
        if _row_has_data(ws, mid, check_cols):
            lo = mid
        else:
            hi = mid

    # Phase 4: Linear vom hi-Punkt zurücklaufen für exakte Grenze
    last_real = lo
    for r in range(lo, min(hi + 1, max_row + 1)):
        if _row_has_data(ws, r, check_cols):
            last_real = r

    return last_real


def _find_last_data_col(ws, max_col: int, sample_rows: int) -> int:
    """Findet die letzte befüllte Spalte per Stichprobe der ersten N Zeilen.

    Strategie: Die Spaltenbreite ist fast immer in den ersten Zeilen
    (Header) erkennbar. Wir prüfen die ersten sample_rows Zeilen und
    finden die am weitesten rechts befüllte Spalte.
    """
    check_rows = min(sample_rows, 50)
    last_col = 0

    # Linear durch die ersten Zeilen, aber exponentiell durch die Spalten
    for r in range(1, check_rows + 1):
        # Erst die bekannten benutzten Spalten prüfen
        for c in range(1, min(last_col + 20, max_col + 1)):
            if ws.cell(row=r, column=c).value is not None:
                last_col = max(last_col, c)
        # Dann exponentiell weiter suchen
        probe_c = last_col + 20
        while probe_c <= max_col:
            if ws.cell(row=r, column=probe_c).value is not None:
                last_col = max(last_col, probe_c)
                probe_c += 10
            else:
                break
            probe_c *= 2

    return last_col


def _count_cells_sampled(ws, real_rows: int, real_cols: int) -> tuple[int, int]:
    """Zählt Formeln und statische Zellen per repräsentativer Stichprobe.

    Strategie: Erste 100 Zeilen vollständig + danach jede 50ste Zeile.
    Hochrechnung basiert auf dem tatsächlichen Datenbereich.
    """
    formula_count = 0
    static_count = 0
    sampled_rows = 0

    for row_idx in range(1, real_rows + 1):
        # Erste 100 Zeilen immer, danach jede 50ste
        if row_idx > 100 and (row_idx % 50) != 0:
            continue
        sampled_rows += 1
        for col_idx in range(1, real_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue
            if cell.data_type == 'f':
                formula_count += 1
            else:
                static_count += 1

    # Hochrechnung: sampled_rows zu real_rows
    if sampled_rows > 0 and sampled_rows < real_rows:
        factor = real_rows / sampled_rows
        formula_count = int(formula_count * factor)
        static_count = int(static_count * factor)

    return formula_count, static_count


# Typ für Progress-Callbacks
ProgressEvent = dict  # {"step": str, "detail": str, "pct": int, "status": "running"|"done"|"error"}


def _compute_db_readiness(stats: SheetStats, findings: list) -> int:
    """Berechnet wie gut ein Sheet in eine Datenbank überführbar wäre (0-100)."""
    score = 100
    name = stats.name
    rules_found = {f.rule_id for f in findings if f.sheet == name}

    # Penalty pro erkanntem Strukturproblem
    if 'STR-006' in rules_found: score -= 25   # Kein Primärschlüssel
    if 'STR-003' in rules_found: score -= 20   # Keine Kopfzeile
    if 'STR-001' in rules_found: score -= 15   # Merged Cells
    if 'STR-002' in rules_found: score -= 10   # Gemischte Datentypen
    if 'STR-005' in rules_found: score -= 10   # Inkonsistente IDs
    if 'STR-007' in rules_found: score -= 10   # Freitext-IDs
    if 'STR-004' in rules_found: score -= 5    # Leerzeilen-Gaps
    if 'FRM-004' in rules_found: score -= 15   # Zirkelbezüge
    if 'FRM-005' in rules_found: score -= 5    # Externe Verknüpfungen

    # Merged Cells auch ohne Finding berücksichtigen (Light-Modus)
    if stats.merged_regions > 0 and 'STR-001' not in rules_found:
        score -= min(15, stats.merged_regions * 3)

    # Sehr kleine Sheets (Konfiguration/Lookup) – schwer bewertbar
    if stats.row_count < 3:
        score = min(score, 50)

    # Formel-dominante Sheets sind Berechnungs-Sheets, nicht Datenspeicher
    total_cells = stats.formula_count + stats.static_count
    if total_cells > 0 and stats.formula_count / total_cells > 0.7:
        score -= 10

    return max(0, min(100, score))


def analyze(
    file_path: str,
    rules: Optional[List[type[BaseRule]]] = None,
) -> WorkbookReport:
    """Analysiert eine Excel-Datei und gibt einen vollständigen Report zurück."""
    # Wrapper um analyze_with_progress, ignoriert Events
    report = None
    for event_or_report in analyze_with_progress(file_path, rules):
        if isinstance(event_or_report, WorkbookReport):
            report = event_or_report
    return report


import threading
import traceback

from excel_checker.rules._sampling import SampleMode

# Tier-Schwellen (MB). Tier 1 = voller Scan, Tier 2 = read-only + Sampling
# für style-lose Regeln, Tier 3 = nur skalierbare Regeln mit aggressiverem
# Sampling. Die Schwellen sind bewusst konservativ gewählt — sie lassen sich
# später experimentell verschieben.
_TIER1_MAX_MB = 15
_TIER2_MAX_MB = 50


def _select_tier(file_size_mb: float) -> tuple[int, Optional[SampleMode], dict]:
    """Wählt Scan-Tier, Sampling-Strategie und openpyxl-Load-Kwargs.

    Rückgabe:
        (tier_nr, sample_mode_or_None, load_kwargs)

    tier 1: alle Regeln, vollständiger Scan, keine Sampling-Einschränkung.
    tier 2: Regeln ohne ``needs_styles`` laufen, sampling-fähige Regeln
            bekommen ``SampleMode`` übergeben. Workbook wird NICHT im
            read-only-Modus geladen, da bestehende Regeln ``ws.cell(r,c)``
            nutzen – das ist in openpyxl read-only ineffizient.
    tier 3: wie Tier 2, aber Regeln die weder sampling-fähig noch
            zellen-unabhängig sind, werden übersprungen. Aggressiveres
            Sampling (kleinere Stichprobe pro Blatt).

    Der read-only-Modus bleibt eine spätere Optimierung (würde erfordern,
    dass Regeln ``iter_rows`` statt ``ws.cell`` nutzen).
    """
    base_load = {"read_only": False, "data_only": False}
    if file_size_mb <= _TIER1_MAX_MB:
        return 1, None, base_load
    if file_size_mb <= _TIER2_MAX_MB:
        return (
            2,
            SampleMode(max_cells_per_sheet=10_000, max_rows_per_sheet=20_000),
            base_load,
        )
    return (
        3,
        SampleMode(max_cells_per_sheet=3_000, max_rows_per_sheet=8_000),
        base_load,
    )


def _filter_rules_for_tier(
    rules: List[type[BaseRule]], tier: int
) -> List[type[BaseRule]]:
    """Filtert Regeln passend zum Tier.

    Tier 1: alle Regeln
    Tier 2: alle Regeln ohne ``needs_styles``
    Tier 3: alle Regeln ohne ``needs_styles`` und, wenn ``scales_with_cells``,
            nur mit ``supports_sampling``
    """
    if tier == 1:
        return list(rules)
    if tier == 2:
        return [r for r in rules if not r.needs_styles]
    # tier 3
    return [
        r
        for r in rules
        if not r.needs_styles
        and (not r.scales_with_cells or r.supports_sampling)
    ]


def analyze_with_progress(
    file_path: str,
    rules: Optional[List[type[BaseRule]]] = None,
    rule_timeout_sec: int = 30,
    progress_callback: Optional[Callable[[dict], None]] = None,
) -> Generator[ProgressEvent | WorkbookReport, None, None]:
    """Analysiert eine Excel-Datei und yielded Progress-Events.

    Das letzte yield ist der fertige WorkbookReport.
    """
    if rules is None:
        rules = ALL_RULES

    total_steps = len(rules) + 3  # +3 für: Laden, Sheet-Stats, Empfehlungen
    current_step = 0

    def progress(step: str, detail: str = "", status: str = "running") -> ProgressEvent:
        nonlocal current_step
        current_step += 1
        pct = min(100, int(current_step / total_steps * 100))
        return {
            "step": step, "detail": detail,
            "pct": pct, "status": status,
            "current": current_step, "total": total_steps,
        }

    def sub_progress(step: str, detail: str = "") -> ProgressEvent:
        """Progress-Update ohne Step-Counter zu erhöhen (für Zwischenstatus)."""
        pct = min(100, int(current_step / total_steps * 100))
        return {
            "step": step, "detail": detail,
            "pct": pct, "status": "running",
            "current": current_step, "total": total_steps,
        }

    file_path = os.path.abspath(file_path)
    if not os.path.exists(file_path):
        raise FileNotFoundError(_("engine.file_not_found", path=file_path))

    file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
    size_str = f"{file_size_mb:.1f} MB" if file_size_mb >= 0.1 else f"{file_size_mb * 1024:.0f} KB"

    # Gestuftes Scan-Modell: siehe _select_tier für die Entscheidungslogik.
    tier, sample_mode, load_kwargs = _select_tier(file_size_mb)
    active_rules = _filter_rules_for_tier(rules, tier)
    # Das alte Verhalten ("keine Regeln bei >15 MB") bleibt als Notanker in
    # tier 2/3 erhalten, falls alle Regeln rausgefiltert werden — die reinen
    # Volumen-Findings werden unten im ``tier >= 2``-Branch generiert.
    large_file = tier >= 2

    total_steps = len(active_rules) + 3  # Laden, Stats, Empfehlungen

    # Vorschau der anstehenden Schritte senden
    queue_items = [_("engine.stats_queue")]
    if large_file:
        queue_items.append(_("engine.volume_light_queue"))
    else:
        rule_names = []
        for rule_cls in active_rules:
            try:
                name = rule_cls().rule_name
            except Exception as e:
                name = f"[Fehler: {e}]"
            print(f"[DEBUG] Regel-Name für {rule_cls.__name__}: {repr(name)}")
            rule_names.append(name)
        queue_items += [_("engine.rule_queue", name=name) for name in rule_names]
    queue_items.append(_("engine.recs_queue"))

    mode_hint = _("engine.light_hint") if large_file else ""
    yield {
        "step": _("engine.loading", mode=mode_hint),
        "detail": _("engine.loading_detail", size=size_str),
        "pct": 0, "status": "running",
        "current": 0, "total": total_steps,
        "queue": queue_items,
    }

    # === Schritt 1: Datei laden ===
    # load_kwargs ist vom Tier bestimmt (read_only=True ab Tier 2).
    wb = openpyxl.load_workbook(file_path, **load_kwargs)
    current_step = 1

    report = WorkbookReport(
        file_path=file_path,
        file_size_mb=file_size_mb,
        sheet_count=len(wb.sheetnames),
    )

    # === Schritt 2: Sheet-Statistiken ===
    sheet_names = wb.sheetnames
    names_str = ', '.join(sheet_names[:5]) + ('...' if len(sheet_names) > 5 else '')
    yield progress(
        _("engine.analyzing"),
        _("engine.analyzing_detail", count=len(sheet_names), names=names_str)
    )

    if large_file:
        # Read-Only: Statistiken über iter_rows
        for ws_idx, ws in enumerate(wb.worksheets):
            yield sub_progress(
                _("engine.sheet_progress", idx=ws_idx + 1, count=len(sheet_names), title=ws.title),
                _("engine.sheet_detail")
            )
            stats = SheetStats(name=ws.title)
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            is_phantom = (max_row >= EXCEL_MAX_ROW or max_col >= EXCEL_MAX_COL
                          or max_row >= _SUSPICIOUS_ROW_THRESHOLD
                          or max_col >= _SUSPICIOUS_COL_THRESHOLD)

            if is_phantom:
                # Phantom-Dimensionen: Vollständiger Scan mit Early-Exit
                stats.is_phantom = True
                scan_cols = min(max_col, 500)
                last_data_row = 0
                formula_count = 0
                static_count = 0
                used_cols = set()
                for row_idx, row in enumerate(ws.iter_rows(max_col=scan_cols), 1):
                    row_has_data = False
                    for cell in row:
                        if cell.value is not None:
                            row_has_data = True
                            used_cols.add(cell.column)
                            if cell.data_type == 'f':
                                formula_count += 1
                            else:
                                static_count += 1
                    if row_has_data:
                        last_data_row = row_idx
                    elif row_idx > last_data_row + _EMPTY_RUN_EXIT and last_data_row > 0:
                        break
                    if last_data_row == 0 and row_idx > 1000:
                        break  # Sheet scheint leer zu sein
                stats.row_count = last_data_row
                stats.col_count = len(used_cols)
                stats.formula_count = formula_count
                stats.static_count = static_count
            else:
                # Kleines Sheet in Read-Only: vollständig scannen
                formula_count = 0
                static_count = 0
                non_empty_rows = 0
                used_cols = set()
                for row in ws.iter_rows(max_row=max_row, max_col=max_col):
                    row_has_data = False
                    for cell in row:
                        if cell.value is None:
                            continue
                        row_has_data = True
                        used_cols.add(cell.column)
                        if cell.data_type == 'f':
                            formula_count += 1
                        else:
                            static_count += 1
                    if row_has_data:
                        non_empty_rows += 1

                stats.row_count = non_empty_rows
                stats.col_count = len(used_cols)
                stats.formula_count = formula_count
                stats.static_count = static_count

            report.sheet_stats.append(stats)

        # Light-Modus: Volumen-basierte Findings generieren
        from excel_checker.models import Finding, Severity, Category
        yield progress(_("engine.volume_light"), _("engine.volume_light_detail"))

        total_rows = sum(s.row_count for s in report.sheet_stats)
        total_formulas = sum(s.formula_count for s in report.sheet_stats)

        if file_size_mb > 50:
            report.findings.append(Finding(
                rule_id="VOL-001", rule_name=_("VOL-001.name"),
                severity=Severity.CRITICAL,
                category=Category.VOLUME,
                message=_("engine.vol_extreme_msg", size=file_size_mb),
                detail=_("engine.vol_extreme_detail"),
                suggestion=_("engine.vol_extreme_tip"),
            ))
        elif file_size_mb > 20:
            report.findings.append(Finding(
                rule_id="VOL-001", rule_name=_("VOL-001.name"),
                severity=Severity.ERROR,
                category=Category.VOLUME,
                message=_("engine.vol_large_msg", size=file_size_mb),
                detail=_("engine.vol_large_detail"),
                suggestion=_("engine.vol_large_tip"),
            ))

        if total_rows > 100000:
            report.findings.append(Finding(
                rule_id="VOL-001", rule_name=_("VOL-001.name"),
                severity=Severity.ERROR,
                category=Category.VOLUME,
                message=_("engine.vol_rows_abused", rows=f"{total_rows:,}"),
                suggestion=_("engine.vol_rows_abused_tip"),
            ))
        elif total_rows > 10000:
            report.findings.append(Finding(
                rule_id="VOL-001", rule_name=_("VOL-001.name"),
                severity=Severity.WARNING,
                category=Category.VOLUME,
                message=_("engine.vol_rows_high", rows=f"{total_rows:,}"),
                suggestion=_("engine.vol_rows_high_tip"),
            ))

        if report.sheet_count > 20:
            report.findings.append(Finding(
                rule_id="VOL-001", rule_name=_("VOL-001.name"),
                severity=Severity.WARNING,
                category=Category.VOLUME,
                message=_("engine.vol_many_sheets", count=report.sheet_count),
                suggestion=_("engine.vol_many_sheets_tip"),
            ))

        report.findings.append(Finding(
            rule_id="VOL-001", rule_name=_("VOL-001.name"),
            severity=Severity.INFO,
            category=Category.VOLUME,
            message=_("engine.light_msg", size=file_size_mb),
            detail=_("engine.light_detail"),
            suggestion=_("engine.light_tip"),
        ))

    else:
        # Normaler Modus: pro Sheet analysieren
        for ws_idx, ws in enumerate(wb.worksheets):
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            is_phantom = (max_row >= EXCEL_MAX_ROW or max_col >= EXCEL_MAX_COL
                          or max_row >= _SUSPICIOUS_ROW_THRESHOLD
                          or max_col >= _SUSPICIOUS_COL_THRESHOLD)

            stats = SheetStats(name=ws.title)
            stats.merged_regions = len(list(ws.merged_cells.ranges))

            if is_phantom:
                # Phantom-Dimensionen: Vollständiger Scan mit Early-Exit
                stats.is_phantom = True
                scan_cols = min(max_col, 500)
                last_data_row = 0
                formula_count = 0
                static_count = 0
                used_cols = set()
                row_idx = 0
                while True:
                    row_idx += 1
                    row_has_data = False
                    for col_idx in range(1, scan_cols + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            row_has_data = True
                            used_cols.add(col_idx)
                            if cell.data_type == 'f':
                                formula_count += 1
                            else:
                                static_count += 1
                    if row_has_data:
                        last_data_row = row_idx
                    elif row_idx > last_data_row + _EMPTY_RUN_EXIT and last_data_row > 0:
                        break
                    if last_data_row == 0 and row_idx > 1000:
                        break  # Sheet scheint leer zu sein
                stats.row_count = last_data_row
                stats.col_count = len(used_cols)
                stats.formula_count = formula_count
                stats.static_count = static_count
            else:
                # Kein Phantom & kein verdächtiger Bereich: Fuzzy-Scan
                # Erst echte Grenzen finden, dann Zellen zählen
                real_cols = _find_last_data_col(ws, max_col, sample_rows=50)
                real_rows = _find_last_data_row(ws, max_row, check_cols=real_cols or max_col)

                if real_rows == 0:
                    real_rows = max_row
                if real_cols == 0:
                    real_cols = max_col

                formula_count, static_count = _count_cells_sampled(
                    ws, real_rows, real_cols
                )

                stats.row_count = real_rows
                stats.col_count = real_cols
                stats.formula_count = formula_count
                stats.static_count = static_count

            report.sheet_stats.append(stats)

    # === Schritt 3-N: Regeln ausführen (nur im Normalmodus) ===
    for rule_idx, rule_cls in enumerate(active_rules):
        rule = rule_cls()
        yield progress(
            _( "engine.rule_checking", name=rule.rule_name),
            _( "engine.rule_detail", idx=rule_idx + 1, count=len(active_rules), rule_id=rule.rule_id)
        )

        findings = []
        error = None
        error_tb = None
        # Sample-Mode nur an sampling-fähige Regeln übergeben. Andere Regeln
        # ignorieren die Stichproben-Schwelle und scannen ihre üblichen
        # Teil-Bereiche (z. B. erste 5000 Zeilen).
        effective_sample = sample_mode if rule_cls.supports_sampling else None

        def run_rule():
            nonlocal findings, error, error_tb
            try:
                # argcount enthält self → 4 bedeutet: check(self, wb, file_path, progress_callback)
                # 5 bedeutet: check(self, wb, file_path, progress_callback, sample_mode)
                argcount = rule.check.__code__.co_argcount
                progress_cb = lambda evt: (progress_callback(evt) if progress_callback else None)
                if argcount >= 5:
                    findings = rule.check(wb, file_path, progress_cb, effective_sample)
                elif argcount >= 4:
                    findings = rule.check(wb, file_path, progress_cb)
                else:
                    findings = rule.check(wb, file_path)
            except Exception as e:
                error = e
                # Traceback JETZT einfangen — format_exc() außerhalb des except-
                # Blocks liefert "NoneType: None\n".
                error_tb = traceback.format_exc()
        thread = threading.Thread(target=run_rule)
        thread.start()
        thread.join(timeout=rule_timeout_sec)
        if thread.is_alive():
            # Timeout: Thread läuft noch. Error-Event darf den Step-Counter
            # NICHT weiter erhöhen — sonst erreicht pct 100 % bevor die letzten
            # Regeln durch sind und die UI zeigt fälschlich "fertig".
            error_evt = sub_progress(
                _( "engine.rule_error", name=rule.rule_name),
                _( "engine.rule_error_detail", error=f"Timeout nach {rule_timeout_sec}s"),
            )
            error_evt["status"] = "error"
            yield error_evt
            continue
        if error:
            error_evt = sub_progress(
                _( "engine.rule_error", name=rule.rule_name),
                _( "engine.rule_error_detail", error=f"{error}\n{error_tb or ''}"),
            )
            error_evt["status"] = "error"
            yield error_evt
            continue
        report.findings.extend(findings)
        # Phase 2 A.3: Regeln können zusätzlich Column-Profiles liefern.
        extra_profiles = getattr(rule, "column_profiles", None)
        if extra_profiles:
            report.column_profiles.extend(extra_profiles)

    # === DB-Readiness pro Sheet berechnen ===
    for stats in report.sheet_stats:
        stats.db_readiness = _compute_db_readiness(stats, report.findings)

    # === Letzter Schritt: Empfehlungen ===
    yield progress(_("engine.recs_progress"), "", "running")
    report.recommendations = generate_recommendations(report)

    # Fertig-Event
    yield {
        "step": _("engine.done"),
        "detail": _("engine.done_detail", score=report.health_score, findings=len(report.findings)),
        "pct": 100, "status": "done",
        "current": total_steps, "total": total_steps,
    }

    wb.close()
    yield report
