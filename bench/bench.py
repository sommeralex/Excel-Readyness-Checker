import sys, time, resource, os
mode, path = sys.argv[1], sys.argv[2]
t0 = time.perf_counter()
if mode in ("load", "load_ro"):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=(mode=="load_ro"), data_only=False)
    ws = wb[wb.sheetnames[0]]
    n = ws.max_row
    if mode == "load_ro":                      # read_only: Zellen wirklich anfassen
        n = sum(1 for _ in ws.iter_rows(max_row=50000))
elif mode == "analyze":
    from excel_checker.engine import analyze
    rep = analyze(path)
    n = f"score={rep.health_score} findings={len(rep.findings)}"
dt = time.perf_counter() - t0
# ru_maxrss ist auf Linux in KB, auf macOS/BSD in Bytes.
_rss_div = 1048576 if sys.platform == "darwin" else 1024
rss = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss / _rss_div
mb = os.path.getsize(path)/1048576
print(f"{mode:9s} {os.path.basename(path):12s} {mb:6.1f} MB  {dt:7.1f} s  {rss:8.0f} MB peak   {n}")
